import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

NAVY    = "FF0A1628"
BLUE    = "FF1E3A5F"
ORANGE  = "FFF47920"
RED     = "FFDC2626"
AMBER   = "FFD97706"
GREEN   = "FF16A34A"
WHITE   = "FFFFFFFF"
LIGHT   = "FFF8FAFC"
LIGHT2  = "FFEEF3F8"
RED_BG  = "FFFEE2E2"
AMB_BG  = "FFFEF3C7"
GRN_BG  = "FFDCFCE7"
DARK    = "FF0A1628"
MID     = "FF111827"
MUTED   = "FF64748B"


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _c(ws, coord, value=None, fill_hex=None, bold=False, size=9,
       font_color=WHITE, halign="center", valign="center", wrap=True):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.font = Font(name="Calibri", bold=bold, size=size, color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    return cell

def _merge(ws, rng, value=None, fill_hex=None, bold=False, size=9,
           font_color=WHITE, halign="left", valign="center", wrap=True):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    return _c(ws, top_left, value, fill_hex, bold, size, font_color, halign, valign, wrap)

def _compliance_color(rate):
    if rate is None:
        return MUTED
    if rate >= 0.85:
        return GREEN
    if rate >= 0.70:
        return AMBER
    return RED

def _ccvs_badge_color(ncr, cond):
    if ncr > 0:
        return RED
    if cond > 0:
        return AMBER
    return GREEN

def _sheet_page_setup(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)


def _build_sheet1(ws, data):
    period  = data["period_label"]
    kpi     = data["kpi"]
    sites   = data["sites"]
    ccvs    = data["ccvs_rows"]
    actions = data["open_actions"]

    col_widths = {
        "A": 1.44140625,
        "B": 12.78, "C": 12.78, "D": 12.78, "E": 12.78, "F": 12.78, "G": 12.78,
        "H": 3.0,   "I": 17.0,  "J": 21.0,
        "K": 9.0,   "L": 9.33203125, "M": 9.6640625,
        "N": 7.44140625, "O": 2.0, "P": 14.0, "Q": 1.44140625
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    for r in range(1, 5):
        ws.row_dimensions[r].height = 10.8
    for r in range(5, 13):
        ws.row_dimensions[r].height = 13.2
    ws.row_dimensions[13].height = 16.05
    for r in range(14, 22):
        ws.row_dimensions[r].height = 13.95
    ws.row_dimensions[21].height = 6.0
    ws.row_dimensions[22].height = 16.05
    ws.row_dimensions[23].height = 13.95
    for r in range(24, 28):
        ws.row_dimensions[r].height = 19.95
    ws.row_dimensions[28].height = 15.0
    ws.row_dimensions[29].height = 13.95

    for r in range(1, 5):
        for col in range(1, 18):
            ws.cell(row=r, column=col).fill = _fill(NAVY)

    _merge(ws, "B2:J3", "PIMS RPD — Manager Report", NAVY, bold=True, size=16, font_color=WHITE, halign="left")
    _merge(ws, "K2:P3", period, NAVY, bold=True, size=20, font_color=ORANGE, halign="right")
    ws["Q2"].fill = _fill(NAVY)
    ws["Q3"].fill = _fill(NAVY)

    for col in ("B", "C"):
        ws[f"{col}6"].fill = _fill(BLUE)
    for col in ("D", "E"):
        ws[f"{col}6"].fill = _fill(RED)
    for col in ("F", "G"):
        ws[f"{col}6"].fill = _fill(AMBER)
    for col in ("P", "Q"):
        ws[f"{col}6"].fill = _fill(GREEN)
    for col, label in [("I","Metric"),("J","Current"),("K","Previous"),("L","Delta"),("M","Status")]:
        _c(ws, f"{col}6", label, BLUE, bold=True, size=9, font_color=WHITE)

    tot    = kpi["total_obs"]
    prev_t = kpi.get("prev_total_obs", 0)
    open_a = kpi["open_actions"]
    prev_o = kpi.get("prev_open_actions", 0)
    ncr    = kpi["ncr"]
    prev_n = kpi.get("prev_ncr", 0)
    cond   = kpi["conditional"]
    prev_c = kpi.get("prev_conditional", 0)
    rate   = kpi["compliance_rate"]
    prev_r = kpi.get("prev_compliance_rate", 0.0)

    _c(ws, "B7", "Total observations", LIGHT,  bold=False, size=8, font_color=MUTED, halign="left")
    _c(ws, "C7", "Open actions",       LIGHT2, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "D7:E7", "NCR",         RED_BG, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "F7:G7", "Conditional", AMB_BG, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "P7:Q7", "Compliance rate", GRN_BG, bold=False, size=8, font_color=MUTED, halign="left")

    _merge(ws, "B8:B9",  tot,    LIGHT,  bold=True, size=20, font_color=BLUE,  halign="left")
    _merge(ws, "C8:C9",  open_a, LIGHT2, bold=True, size=20, font_color=BLUE,  halign="left")
    _merge(ws, "D8:E9",  ncr,    RED_BG, bold=True, size=20, font_color=RED,   halign="left")
    _merge(ws, "F8:G9",  cond,   AMB_BG, bold=True, size=20, font_color=AMBER, halign="left")
    _merge(ws, "P8:Q9",  rate,   GRN_BG, bold=True, size=20, font_color=GREEN, halign="left")
    ws["P8"].number_format = "0%"

    _c(ws, "B10", f"+{tot - prev_t} vs prev month",    LIGHT,  bold=False, size=8, font_color=MUTED, halign="left")
    _c(ws, "C10", f"+{open_a - prev_o} vs prev month", LIGHT2, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "D10:E10", f"+{ncr - prev_n} vs prev month",  RED_BG, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "F10:G10", f"+{cond - prev_c} vs prev month", AMB_BG, bold=False, size=8, font_color=MUTED, halign="left")
    _merge(ws, "P10:Q10", "Target: 85%", GRN_BG, bold=False, size=8, font_color=MUTED, halign="left")

    kpi_rows = [
        ("Total observations", tot,           prev_t,  f"+{tot-prev_t}",           "up"),
        ("Compliance rate",    f"{rate:.0%}", f"{prev_r:.0%}", f"+{(rate-prev_r):.0%}", "up"),
        ("NCR",                ncr,           prev_n,  f"+{ncr-prev_n}",            "down"),
        ("Conditional",        cond,          prev_c,  f"+{cond-prev_c}",           "down"),
        ("Open actions",       open_a,        prev_o,  f"+{open_a-prev_o}",         "down"),
    ]
    row_fills     = [LIGHT, WHITE, LIGHT, WHITE, LIGHT]
    status_colors = [GREEN, GREEN, RED, AMBER, RED]

    for i, (metric, cur, prev, delta, good_dir) in enumerate(kpi_rows):
        r  = 7 + i
        bg = row_fills[i]
        try:
            d_val = float(str(delta).replace("+","").replace("%",""))
            if "%" in str(delta):
                arrow_txt = f"▲ +{str(delta).lstrip('+')}" if d_val >= 0 else f"▼ {delta}"
            else:
                arrow_txt = f"▲ +{int(d_val)}" if d_val >= 0 else f"▼ {int(d_val)}"
        except:
            arrow_txt = delta
        _c(ws, f"I{r}", metric,    bg, bold=True,  size=10, font_color=DARK, halign="left")
        _c(ws, f"J{r}", cur,       bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"K{r}", prev,      bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"L{r}", delta,     bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"M{r}", arrow_txt, bg, bold=True,  size=10, font_color=status_colors[i], halign="center")

    _merge(ws, "B13:G13", "Site Breakdown",               BLUE, bold=True, size=9, halign="left")
    _merge(ws, "I13:N13", "CCVS Monitoring & Measurement", BLUE, bold=True, size=9, halign="left")
    _merge(ws, "P13:P14", "Compliance", BLUE, bold=True, size=9, halign="center", valign="center")

    _merge(ws, "B14:C14", "Site Address", BLUE, bold=True, size=9, halign="center")
    for col, lbl in [("D","NCR"),("E","Conditional"),("F","Compliant"),("G","Total")]:
        _c(ws, f"{col}14", lbl, BLUE, bold=True, size=9)
    for col, lbl in [("I","Code"),("J","Category"),("K","NCR"),("L","Cond"),("M","Comp"),("N","Total")]:
        _c(ws, f"{col}14", lbl, BLUE, bold=True, size=9)

    site = sites[0] if sites else {"address": "", "ncr": 0, "conditional": 0, "compliant": 0, "total": 0}
    site_comp_pct = (site["compliant"] / site["total"]) if site["total"] else 0.0
    _merge(ws, "B15:C15", site.get("address",""), LIGHT, bold=False, size=9, font_color=MID, halign="left")
    _c(ws, "D15", site["ncr"],         RED_BG, bold=True, size=11, font_color=RED,   halign="center")
    _c(ws, "E15", site["conditional"], AMB_BG, bold=True, size=11, font_color=AMBER, halign="center")
    _c(ws, "F15", site["compliant"],   GRN_BG, bold=True, size=11, font_color=GREEN, halign="center")
    _c(ws, "G15", site["total"],       LIGHT,  bold=True, size=11, font_color=BLUE,  halign="center")
    _c(ws, "P15", f"{int(site_comp_pct*100)}%", LIGHT, bold=True, size=10,
       font_color=_compliance_color(site_comp_pct), halign="center")

    ccvs_bg = [LIGHT, WHITE, LIGHT, WHITE, LIGHT, WHITE]
    for i, row_data in enumerate(ccvs[:6]):
        r  = 15 + i
        bg = ccvs_bg[i]
        badge_fill = _ccvs_badge_color(row_data["ncr"], row_data["conditional"])
        _c(ws, f"I{r}", row_data["ccvs_code"], badge_fill, bold=True,  size=9, font_color=WHITE, halign="center")
        _c(ws, f"J{r}", row_data["category"], bg,         bold=False, size=9, font_color=MID,   halign="left")
        _c(ws, f"K{r}", row_data["ncr"],         bg, bold=False, size=9, font_color=MID, halign="center")
        _c(ws, f"L{r}", row_data["conditional"], bg, bold=False, size=9, font_color=MID, halign="center")
        _c(ws, f"M{r}", row_data["compliant"],   bg, bold=False, size=9, font_color=MID, halign="center")
        _c(ws, f"N{r}", row_data["total"],       bg, bold=False, size=9, font_color=MID, halign="center")
        ccvs_pct = row_data.get("compliance_rate", 0) / 100
        _c(ws, f"P{r}", f"{int(ccvs_pct*100)}%", bg, bold=True, size=10,
           font_color=_compliance_color(ccvs_pct), halign="center")

    open_count = len(actions)
    _merge(ws, "B22:P22", f"Open Actions Register  ({open_count} open)", BLUE, bold=True, size=9, halign="left")
    _c(ws, "B23", "#",      NAVY, bold=True, size=9)
    _c(ws, "C23", "CCVS",   NAVY, bold=True, size=9)
    _c(ws, "D23", "Status", NAVY, bold=True, size=9)
    _merge(ws, "E23:L23", "Action Required", NAVY, bold=True, size=9, halign="left")
    _merge(ws, "M23:N23", "Responsible",     NAVY, bold=True, size=9)
    _c(ws, "P23", "Due",    NAVY, bold=True, size=9)

    for i, act in enumerate(actions[:4]):
        r      = 24 + i
        status = act.get("status", "NCR")
        if status == "NCR":
            row_bg, badge_bg, due_col = RED_BG, RED, RED
        elif status in ("Conditional", "Cond"):
            row_bg, badge_bg, due_col = AMB_BG, AMBER, AMBER
        else:
            row_bg, badge_bg, due_col = GRN_BG, GREEN, GREEN
        _c(ws, f"B{r}", str(act.get("obs_id","")),        row_bg,   bold=True,  size=9, font_color=DARK, halign="center")
        _c(ws, f"C{r}", act.get("ccvs_code",""),           badge_bg, bold=True,  size=9, font_color=WHITE)
        _c(ws, f"D{r}", status,                             badge_bg, bold=True,  size=9, font_color=WHITE)
        _merge(ws, f"E{r}:L{r}", act.get("action_text",""),  row_bg, bold=False, size=9, font_color=MID, halign="left")
        _merge(ws, f"M{r}:N{r}", act.get("responsible",""),   row_bg, bold=False, size=9, font_color=MID, halign="center")
        _c(ws, f"P{r}", act.get("due",""),                 row_bg,   bold=True,  size=9, font_color=due_col, halign="center")

    _merge(ws, "B29:P29", "Live audit records only  ·  staging=FALSE AND source_pdf=FALSE",
           None, bold=False, size=8, font_color=MUTED, halign="left")
    _sheet_page_setup(ws)


def _build_sheet2(ws, data):
    period  = data["period_label"]
    actions = data["open_actions"]

    for col, w in {"A": 1.44, "B": 5.0, "C": 10.0, "D": 11.0, "E": 10.0,
                   "F": 36.0, "G": 14.0, "H": 12.0, "I": 28.0, "J": 28.0, "K": 1.44}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 6.0
    ws.row_dimensions[2].height = 24.0
    ws.row_dimensions[3].height = 18.0
    ws.row_dimensions[4].height = 6.0
    ws.row_dimensions[5].height = 4.95
    ws.row_dimensions[6].height = 16.05

    for r in range(1, 5):
        for col in range(1, 12):
            ws.cell(row=r, column=col).fill = _fill(NAVY)

    open_count = len(actions)
    _merge(ws, "B2:H3", f"Open Actions Register — {period}", NAVY, bold=True, size=14, font_color=WHITE, halign="left")
    _merge(ws, "I2:J3", f"{open_count} open", NAVY, bold=True, size=13, font_color=RED, halign="right")
    ws["K2"].fill = _fill(NAVY)
    ws["K3"].fill = _fill(NAVY)

    for col, lbl in [("B","#"),("C","Site"),("D","Date"),("E","Status"),("F","Action Required"),
                     ("G","Responsible"),("H","Due"),("I","Monitoring Note"),("J","Observation")]:
        _c(ws, f"{col}6", lbl, BLUE, bold=True, size=9)

    for i, act in enumerate(actions):
        r      = 7 + i
        ws.row_dimensions[r].height = 36.0
        status = act.get("status", "NCR")
        if status == "NCR":
            row_bg, badge_bg, due_col = RED_BG, RED, RED
        elif status in ("Conditional", "Cond"):
            row_bg, badge_bg, due_col = AMB_BG, AMBER, AMBER
        else:
            row_bg, badge_bg, due_col = GRN_BG, GREEN, GREEN
        _c(ws, f"B{r}", str(act.get("obs_id","")),        row_bg,   bold=True,  size=9, font_color=DARK, halign="center")
        _c(ws, f"C{r}", act.get("site",""),                row_bg,   bold=False, size=9, font_color=MID,  halign="left")
        _c(ws, f"D{r}", act.get("date",""),                row_bg,   bold=False, size=9, font_color=MID,  halign="center")
        _c(ws, f"E{r}", status,                             badge_bg, bold=True,  size=9, font_color=WHITE)
        _c(ws, f"F{r}", act.get("action_text",""),         row_bg,   bold=False, size=9, font_color=MID,  halign="left")
        _c(ws, f"G{r}", act.get("responsible",""),         row_bg,   bold=False, size=9, font_color=MID,  halign="center")
        _c(ws, f"H{r}", act.get("due",""),                 row_bg,   bold=True,  size=9, font_color=due_col, halign="center")
        _c(ws, f"I{r}", act.get("monitoring_note",""),     row_bg,   bold=False, size=9, font_color=MID,  halign="left")
        _c(ws, f"J{r}", act.get("observation",""),         row_bg,   bold=False, size=9, font_color=MID,  halign="left")

    _sheet_page_setup(ws)


def _build_sheet3(ws, data):
    period = data["period_label"]
    kpi    = data["kpi"]

    for col, w in {"A": 1.44, "B": 28.0, "C": 14.0, "D": 12.0,
                   "E": 12.0, "F": 16.0, "G": 1.44}.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 6.0
    ws.row_dimensions[2].height = 24.0
    ws.row_dimensions[3].height = 18.0
    ws.row_dimensions[4].height = 6.0
    ws.row_dimensions[5].height = 4.95
    ws.row_dimensions[6].height = 16.05
    for r in range(7, 12):
        ws.row_dimensions[r].height = 18.0
    ws.row_dimensions[12].height = 7.95
    ws.row_dimensions[13].height = 15.0

    for r in range(1, 5):
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = _fill(NAVY)
    _merge(ws, "B2:F3", f"KPI Summary — {period}", NAVY, bold=True, size=14, font_color=WHITE, halign="left")
    ws["G2"].fill = _fill(NAVY)
    ws["G3"].fill = _fill(NAVY)

    for col, lbl in [("B","Metric"),("C","Current"),("D","Previous"),("E","Delta"),("F","Status")]:
        _c(ws, f"{col}6", lbl, BLUE, bold=True, size=9)

    tot    = kpi["total_obs"]
    prev_t = kpi.get("prev_total_obs", 0)
    open_a = kpi["open_actions"]
    prev_o = kpi.get("prev_open_actions", 0)
    ncr    = kpi["ncr"]
    prev_n = kpi.get("prev_ncr", 0)
    cond   = kpi["conditional"]
    prev_c = kpi.get("prev_conditional", 0)
    rate   = kpi["compliance_rate"]
    prev_r = kpi.get("prev_compliance_rate", 0.0)

    rows = [
        ("Total observations", tot,           prev_t,  f"+{tot-prev_t}",           f"▲ +{tot-prev_t}",           GREEN),
        ("Compliance rate",    f"{rate:.0%}", f"{prev_r:.0%}", f"+{(rate-prev_r):.0%}", f"▲ +{(rate-prev_r):.0%}", GREEN),
        ("NCR",                ncr,           prev_n,  f"+{ncr-prev_n}",            f"▲ +{ncr-prev_n}",            RED),
        ("Conditional",        cond,          prev_c,  f"+{cond-prev_c}",           f"▲ +{cond-prev_c}",           AMBER),
        ("Open actions",       open_a,        prev_o,  f"+{open_a-prev_o}",         f"▲ +{open_a-prev_o}",         RED),
    ]
    row_bgs = [LIGHT, WHITE, LIGHT, WHITE, LIGHT]
    for i, (metric, cur, prev, delta, arrow, s_color) in enumerate(rows):
        r = 7 + i
        bg = row_bgs[i]
        _c(ws, f"B{r}", metric, bg, bold=True,  size=10, font_color=DARK, halign="left")
        _c(ws, f"C{r}", cur,    bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"D{r}", prev,   bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"E{r}", delta,  bg, bold=False, size=10, font_color=MID,  halign="center")
        _c(ws, f"F{r}", arrow,  bg, bold=True,  size=10, font_color=s_color, halign="center")

    _merge(ws, "B13:F13", "Live audit records only  ·  staging=FALSE AND source_pdf=FALSE",
           None, bold=False, size=8, font_color=MUTED, halign="left")
    _sheet_page_setup(ws)


def build_manager_report_xlsx(data: dict) -> bytes:
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Manager Dashboard"
    ws2 = wb.create_sheet("Open Actions")
    ws3 = wb.create_sheet("KPI Summary")
    _build_sheet1(ws1, data)
    _build_sheet2(ws2, data)
    _build_sheet3(ws3, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
