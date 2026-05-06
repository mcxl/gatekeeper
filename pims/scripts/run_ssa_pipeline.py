"""Manual CLI for the SSA evidence-folder → 3-deliverable pipeline.

    python -m pims.scripts.run_ssa_pipeline "<folder>"

Bypasses the watcher; runs the same pipeline once over the supplied
folder. The watcher (added separately) is the long-running entry that
drives this same orchestration on quiescent folders.

v1 scope:
  - parse Evidence_Master.csv, match photos, extract site address
  - lift to EnrichedRow (every row Unmatched until CCVS auto-matcher
    lands as a separate slice)
  - build all three deliverables
  - compute staging_status (tri-state per workflow plan)
  - write sentinel file when applicable
  - write .ssa_run.json with outputs + warnings + status

v1 NON-scope (deferred slices, called out in the plan):
  - input-manifest sha256 + idempotency skip (watcher concern)
  - LLM finding-rewrite + narrative summary (separate async pass)
  - staging 5 MB progressive-downscale + size-based split
  - quiescence polling
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from pims.services.ssa_checklist_lookup import ChecklistLookup
from pims.services.ssa_pipeline import (
    EnrichedRow,
    apply_ra_labels_to_rows,
    build_pims_enriched_xlsx,
    build_pims_staging_xlsx_with_size_control,
    build_ssa_report_docx,
    enrich_observations,
    extract_site_address,
    match_photos,
    parse_evidence_csv,
    parse_merge_argument,
    parse_prior_report_recommendations,
    split_multi_issue_observations,
)

log = logging.getLogger("ssa.cli")


# Folder name contract: YYYY-MM-DD-<CLIENT>[-NN], CLIENT ∈ {RPD, SDG},
# optional "-NN" sub-id distinguishes multiple visits to the same site
# on the same day (e.g. morning + afternoon walk, or sub-area split).
# When present, the sub-id is propagated to the output filenames so
# multiple runs against the same day's folders don't collide.
_FOLDER_RE = re.compile(
    r"^(\d{4})-(\d{2})-(\d{2})-(RPD|SDG)(?:-(\d{2}))?$"
)

# Image extensions the watcher cares about. PNG-with-transparency is
# legal but rare; HEIC explicitly out of scope (filename canonicalisation
# rules in the plan).
_IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

# Client-bulk-endpoint capability (gate 0). RPD has it today, SDG
# doesn't — confirmed against pims/routes.py:2091.
_BULK_ENDPOINT = {"RPD": "/pims/upload/observations", "SDG": None}


def _parse_folder(folder: Path) -> tuple[str, str, str, str, str]:
    """Return (audit_date_iso, audit_date_ddmmyyyy, yymmdd, client, sub_id).

    ``sub_id`` is the trailing ``-NN`` digit pair (e.g. ``"01"``) when
    the folder name carries one, or ``""`` when it doesn't. The sub-id
    is propagated to output filenames so multiple visits on the same
    day (``2026-05-01-SDG-01``, ``...-SDG-02``) don't overwrite each
    other.

    Raises ValueError when the folder name doesn't match the contract.
    """
    m = _FOLDER_RE.match(folder.name)
    if not m:
        raise ValueError(
            f"folder name {folder.name!r} does not match "
            f"YYYY-MM-DD-<CLIENT>[-NN] with CLIENT in (RPD, SDG)"
        )
    yyyy, mm, dd, client, sub_id = m.groups()
    sub_id = sub_id or ""
    iso = f"{yyyy}-{mm}-{dd}"
    ddmmyyyy = f"{dd}/{mm}/{yyyy}"
    yymmdd = f"{yyyy[2:]}{mm}{dd}"
    # Sanity-check the date itself; ValueError on Feb 30 etc.
    datetime.strptime(iso, "%Y-%m-%d")
    return iso, ddmmyyyy, yymmdd, client, sub_id


def _output_names(yymmdd: str, client: str, sub_id: str = "") -> dict[str, str]:
    suffix = f"-{client}-{sub_id}" if sub_id else f"-{client}"
    return {
        "enriched": f"PIMS-Enriched-{yymmdd}{suffix}.xlsx",
        "report": f"Site-Safety-Audit-Report-{yymmdd}{suffix}.docx",
        "staging": f"Site-Visit-Report-Upload-PIMS-Staging-{yymmdd}{suffix}.xlsx",
    }


def _images_in(folder: Path) -> list[Path]:
    return sorted(
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() in _IMAGE_EXTS
    )


# Filename-date-suffix on a prior SSA report:
# ``Site-Safety-Audit-Report-YYMMDD-<CLIENT>[-NN].docx``.
_PRIOR_REPORT_RE = re.compile(
    r"^Site-Safety-Audit-Report-(\d{6})-(RPD|SDG)(?:-\d{2})?\.docx$"
)


def _qualifying_prior_reports(
    folder: Path, current_iso: str, current_target: str,
) -> list[Path]:
    """Return prior SSA reports eligible for input-manifest inclusion.

    Eligible iff ALL:
      - filename matches ``Site-Safety-Audit-Report-YYMMDD-<CLIENT>.docx``
      - parsed date strictly earlier than the current folder's audit date
      - filename != the current target output filename

    Per the plan's "Prior-report reuse policy". Files whose date suffix
    is missing/unparseable are non-qualifying — never guessed from mtime.
    """
    out: list[Path] = []
    for p in folder.iterdir():
        if not p.is_file():
            continue
        if p.name == current_target:
            continue
        m = _PRIOR_REPORT_RE.match(p.name)
        if not m:
            continue
        yymmdd = m.group(1)
        try:
            cand_iso = datetime.strptime(yymmdd, "%y%m%d").date().isoformat()
        except ValueError:
            continue
        if cand_iso < current_iso:
            out.append(p)
    return sorted(out)


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def _compute_manifest(
    csv_path: Path, images: list[Path], prior_reports: list[Path],
) -> str:
    """Full-content sha256 over CSV + images + qualifying prior reports.

    Per the watcher contract: hash the sorted ``filename || sha256(bytes)``
    pairs so that any byte-level change in any input flips the manifest.
    Filenames are lowercased for case-insensitive volumes (Windows).
    """
    parts: list[str] = []
    for p in [csv_path, *images, *prior_reports]:
        parts.append(f"{p.name.lower()}||{_file_sha256(p)}")
    parts.sort()
    h = hashlib.sha256()
    for line in parts:
        h.update(line.encode("utf-8"))
        h.update(b"\n")
    return h.hexdigest()


def _existing_run_record(folder: Path) -> dict | None:
    rj = folder / ".ssa_run.json"
    if not rj.exists():
        return None
    try:
        return json.loads(rj.read_text(encoding="utf-8"))
    except Exception:
        log.warning(".ssa_run.json unreadable; treating as missing")
        return None


def _write_sentinel(folder: Path, name: str, body: str) -> str:
    path = folder / name
    path.write_text(body, encoding="utf-8")
    return path.name


def _resolve_staging_status(
    client: str, site_address: str | None,
) -> tuple[str, str | None]:
    """Return (staging_status, blocker) per the tri-state contract."""
    if not site_address:
        return "not_uploadable", "site_address_unresolved"
    if _BULK_ENDPOINT.get(client) is None:
        return "schema_valid_no_endpoint", None
    return "bulk_uploadable", None


def _apply_vision_enrichment(
    enriched: list[EnrichedRow],
    site_address: str | None,
    audit_date_iso: str,
    enable: bool,
    ra_context: str = "",
) -> tuple[str, dict]:
    """Vision-enabled per-row classification + narrative summary.

    Returns ``(narrative_paragraph, diagnostics_dict)``. On any failure
    path (LLM disabled, key missing, network error, JSON parse error,
    timeout) the function returns an empty narrative and a diagnostics
    dict describing what happened — never raises into the orchestrator.

    Side-effect: in-place mutation of every row that the LLM
    successfully classifies — sets ``conformance_status``,
    ``ccvs_code``, ``ccvs_category``, ``finding``, ``legal_ref``,
    ``recommendation``, ``monitoring_note``. Rows that fall through
    the LLM (no photo / API error / parse error) keep their default
    ``Unmatched`` state.
    """
    if not enable or not enriched:
        return "", {"enabled": False, "rows_total": len(enriched)}

    from pims.services.ssa_vision_enricher import (
        enrich_rows_with_vision,
        generate_narrative_summary,
    )

    async def _drive() -> tuple[str, dict]:
        diag = await enrich_rows_with_vision(
            enriched,
            site_address=site_address or "",
            audit_date_iso=audit_date_iso,
            ra_context=ra_context,
        )
        text = await generate_narrative_summary(
            enriched,
            site_address=site_address or "",
            audit_date_iso=audit_date_iso,
        )
        return text, diag

    try:
        narrative, diag = asyncio.run(_drive())
    except Exception as exc:
        log.warning("vision enrichment driver failed: %s", exc, exc_info=True)
        return "", {
            "enabled": True, "driver_error": f"{type(exc).__name__}: {exc}",
            "rows_total": len(enriched),
        }
    diag["enabled"] = True
    return narrative, diag


class PreflightError(RuntimeError):
    """Raised when a required runtime precondition is missing.

    Distinct from generic ``RuntimeError`` so the CLI can surface the
    failure with a clean exit code and human-readable message before
    any rows are processed (rather than silently producing
    semantically-degraded output).
    """


def _preflight(enrich: bool) -> None:
    """Fail loud BEFORE any row processing when required runtime
    preconditions are missing.

    Currently checks:
      - When ``enrich`` is True, ``ANTHROPIC_API_KEY`` must be set in
        the environment. Without it the vision enricher silently
        leaves every row at ``status="Unmatched"``, and operators
        have hit this twice — once mistaking the result for a code
        bug, once for a billing problem. A loud preflight failure
        prevents both.

    Caller can short-circuit with ``--no-enrich`` when they
    deliberately want the offline path.
    """
    if enrich and not os.environ.get("ANTHROPIC_API_KEY"):
        raise PreflightError(
            "ANTHROPIC_API_KEY is not set in the environment. "
            "Either set the key (load .env, run with the key exported, "
            "or invoke from a shell that already has it) or pass "
            "--no-enrich to run the deterministic offline path "
            "(every row will land Unmatched)."
        )


def _serialise_rows(rows: list[EnrichedRow]) -> list[dict]:
    """Convert EnrichedRow + nested ObservationRow into JSON-safe dicts.

    Used by the two-phase workflow so phase 1 (enrich-only) can
    persist the full row state and phase 2 (from-state) can rebuild
    EnrichedRow objects after the operator has edited the enriched
    xlsx.
    """
    out: list[dict] = []
    for r in rows:
        obs = r.obs
        out.append({
            "obs": {
                "csv_row": obs.csv_row,
                "timestamp_raw": obs.timestamp_raw,
                "timestamp_iso": obs.timestamp_iso,
                "observation_text": obs.observation_text,
                "csv_filename": obs.csv_filename,
                "resolved_filename": obs.resolved_filename,
                "resolved_path": (
                    str(obs.resolved_path) if obs.resolved_path else None
                ),
                "needs_review": obs.needs_review,
                "review_reasons": list(obs.review_reasons),
                "duplicate_filename": obs.duplicate_filename,
            },
            "observation_text_clean": r.observation_text_clean,
            "finding": r.finding,
            "conformance_status": r.conformance_status,
            "ccvs_code": r.ccvs_code,
            "ccvs_category": r.ccvs_category,
            "action_description": r.action_description,
            "recommendation": r.recommendation,
            "legal_ref": r.legal_ref,
            "monitoring_note": r.monitoring_note,
            "location": r.location,
            "hierarchy_of_control": r.hierarchy_of_control,
            "finding_title": r.finding_title,
            "timeframe": r.timeframe,
            "phase": r.phase,
            "activity_ref": r.activity_ref,
            "hold_point": r.hold_point,
            "hrcw": r.hrcw,
            "swms_required": r.swms_required,
            "swms_present": r.swms_present,
            "initial_risk": r.initial_risk,
            "residual_risk": r.residual_risk,
            "evidence_csv_indices": list(r.evidence_csv_indices or []),
        })
    return out


def _deserialise_rows(payload: list[dict]) -> list[EnrichedRow]:
    """Inverse of ``_serialise_rows``. Returns EnrichedRow objects with
    nested ObservationRow rebuilt; ``resolved_path`` is restored as a
    Path when present."""
    from pims.services.ssa_pipeline import EnrichedRow, ObservationRow
    out: list[EnrichedRow] = []
    for d in payload:
        o = d["obs"]
        obs = ObservationRow(
            csv_row=o["csv_row"],
            timestamp_raw=o["timestamp_raw"],
            timestamp_iso=o["timestamp_iso"],
            observation_text=o["observation_text"],
            csv_filename=o["csv_filename"],
            resolved_filename=o.get("resolved_filename"),
            resolved_path=Path(o["resolved_path"]) if o.get("resolved_path") else None,
            needs_review=o.get("needs_review", False),
            review_reasons=list(o.get("review_reasons") or []),
            duplicate_filename=o.get("duplicate_filename", False),
        )
        out.append(EnrichedRow(
            obs=obs,
            observation_text_clean=d.get("observation_text_clean", ""),
            finding=d.get("finding", ""),
            conformance_status=d.get("conformance_status", "Unmatched"),
            ccvs_code=d.get("ccvs_code", ""),
            ccvs_category=d.get("ccvs_category", ""),
            action_description=d.get("action_description", ""),
            recommendation=d.get("recommendation", ""),
            legal_ref=d.get("legal_ref", ""),
            monitoring_note=d.get("monitoring_note", ""),
            location=d.get("location", ""),
            hierarchy_of_control=d.get("hierarchy_of_control", ""),
            finding_title=d.get("finding_title", ""),
            timeframe=d.get("timeframe", ""),
            phase=d.get("phase", ""),
            activity_ref=d.get("activity_ref", ""),
            hold_point=d.get("hold_point", ""),
            hrcw=d.get("hrcw", ""),
            swms_required=bool(d.get("swms_required", False)),
            swms_present=d.get("swms_present", ""),
            initial_risk=d.get("initial_risk", ""),
            residual_risk=d.get("residual_risk", ""),
            evidence_csv_indices=list(d.get("evidence_csv_indices") or []),
        ))
    return out


# Columns in the Enriched Register sheet that the operator is
# expected to edit during a phase-1 review pass. When phase 2 reads
# the (potentially edited) xlsx it overrides the corresponding fields
# on the EnrichedRow loaded from the JSON state.
_ENRICHED_EDITABLE_COLUMNS: dict[str, str] = {
    "observation": "finding",                 # the enriched narrative
    "conformance status": "conformance_status",
    "ccvs code": "ccvs_code",
    "ccvs category": "ccvs_category",
    "action description": "action_description",
    "monitoring note": "monitoring_note",
    "responsible": None,                       # blank-by-default cell
    "due": None,                               # blank-by-default cell
}


def _snapshot_enriched_xlsx(xlsx_path: Path) -> dict[str, dict[str, str]]:
    """Hash every editable cell of every data row keyed by Filename.

    Used by phase 1 to record what was written so phase 2 can tell
    operator edits apart from rendered-fallback values that didn't
    actually change. Returns ``{filename: {header_lc: cell_value}}``.
    """
    if not xlsx_path.exists():
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Enriched Register" not in wb.sheetnames:
        return {}
    ws = wb["Enriched Register"]
    headers = [
        ("" if c.value is None else str(c.value).strip().lower())
        for c in ws[1]
    ]
    try:
        filename_col = headers.index("filename") + 1
    except ValueError:
        return {}
    out: dict[str, dict[str, str]] = {}
    for excel_row in range(2, ws.max_row + 1):
        fn_cell = ws.cell(row=excel_row, column=filename_col).value
        fn = str(fn_cell).strip() if fn_cell else ""
        if not fn:
            continue
        snap: dict[str, str] = {}
        for col_idx, hdr in enumerate(headers, start=1):
            if hdr not in _ENRICHED_EDITABLE_COLUMNS:
                continue
            v = ws.cell(row=excel_row, column=col_idx).value
            snap[hdr] = "" if v is None else str(v).strip()
        out[fn] = snap
    return out


# Per-finding detail-table label → EnrichedRow attribute. Used by
# phase 3 (--from-report) to read operator edits out of the docx
# detail tables and write them back onto EnrichedRows.
_DETAIL_LABEL_TO_ATTR: dict[str, str] = {
    "location": "location",
    "observation": "finding",
    "regulatory basis": "legal_ref",
    "hierarchy of control": "hierarchy_of_control",
    "recommendation": "recommendation",
    "required action": "recommendation",   # legacy templates
    "timeframe": "timeframe",
}


def _extract_detail_table_edits(
    docx_path: Path,
) -> list[dict[str, str]]:
    """Walk the rendered audit report docx and return one dict per
    per-finding detail table (in document order). Each dict maps
    EnrichedRow attribute → cell text. Returns ``[]`` when no detail
    tables are found.
    """
    if not docx_path.exists():
        return []
    from docx import Document
    doc = Document(docx_path)
    out: list[dict[str, str]] = []
    for tbl in doc.tables:
        if len(tbl.columns) != 2 or not tbl.rows:
            continue
        if tbl.rows[0].cells[0].text.strip() != "Location":
            continue
        edit: dict[str, str] = {}
        for row in tbl.rows:
            if len(row.cells) < 2:
                continue
            label = row.cells[0].text.strip().lower()
            attr = _DETAIL_LABEL_TO_ATTR.get(label)
            if attr is None:
                continue
            edit[attr] = row.cells[1].text.strip()
        if edit:
            out.append(edit)
    return out


def _merge_edits_from_audit_report(
    rows: list[EnrichedRow], docx_path: Path,
    finding_render_order: list[int] | None,
) -> dict:
    """Phase 3: read operator edits from the audit report docx's
    per-finding detail tables and apply them back to the matching
    EnrichedRows.

    ``finding_render_order`` is the list of csv-row indices the
    detail tables were rendered FOR (in display order). Phase 2
    stamps it onto the state JSON so phase 3 can pair detail-table
    edits back to the correct EnrichedRow even after merge / sort.

    Returns a diagnostics dict shaped:
        {
          "applied":          bool,
          "tables_seen":      int,
          "tables_paired":    int,
          "field_overrides":  int,
          "unpaired":         int,    # detail tables we couldn't map
        }
    """
    if not docx_path.exists():
        return {"applied": False, "reason": "audit report missing"}
    edits = _extract_detail_table_edits(docx_path)
    if not edits:
        return {"applied": False, "reason": "no detail tables found"}

    # Build csv_idx → EnrichedRow lookup. evidence_csv_indices on a
    # consolidated row include all csv indices folded into it; the
    # canonical row's primary csv_idx is the first entry in that
    # list, which matches what phase 2 emitted.
    by_idx: dict[int, EnrichedRow] = {}
    for r in rows:
        if not r.evidence_csv_indices:
            r.evidence_csv_indices = [r.obs.csv_row]
        by_idx[r.evidence_csv_indices[0]] = r
        # Allow lookup by any folded csv_idx so the operator can
        # also use the un-merged number if they're cross-referencing
        # the original CSV.
        for fold_idx in r.evidence_csv_indices[1:]:
            by_idx.setdefault(fold_idx, r)

    overrides = 0
    paired = 0
    unpaired = 0
    for table_pos, edit in enumerate(edits):
        target = None
        if finding_render_order and table_pos < len(finding_render_order):
            target = by_idx.get(finding_render_order[table_pos])
        if target is None:
            unpaired += 1
            continue
        paired += 1
        for attr, new_val in edit.items():
            old_val = getattr(target, attr, "") or ""
            if (new_val or "") == old_val:
                continue
            setattr(target, attr, new_val or "")
            overrides += 1
    return {
        "applied": True,
        "tables_seen": len(edits),
        "tables_paired": paired,
        "field_overrides": overrides,
        "unpaired": unpaired,
    }


def _merge_edits_from_enriched_xlsx(
    rows: list[EnrichedRow], xlsx_path: Path,
    snapshot: dict[str, dict[str, str]] | None = None,
) -> dict:
    """Read the enriched xlsx and overwrite editable fields on each
    EnrichedRow with the operator's edits.

    Pairing key: the ``Filename`` column (resolved on-disk filename)
    matches each xlsx data row to exactly one EnrichedRow. When the
    filename is ambiguous (post-split composite notes share a
    photo) only the FIRST matching row's editable fields are
    overwritten — the operator should re-split manually if they
    need to differentiate.

    Returns a diagnostics dict counting overrides applied.
    """
    if not xlsx_path.exists():
        return {"applied": False, "reason": "enriched xlsx missing"}
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Enriched Register" not in wb.sheetnames:
        return {"applied": False, "reason": "Enriched Register sheet missing"}
    ws = wb["Enriched Register"]
    headers = [
        ("" if c.value is None else str(c.value).strip().lower())
        for c in ws[1]
    ]
    try:
        filename_col = headers.index("filename") + 1
    except ValueError:
        return {"applied": False, "reason": "filename column missing"}

    # Index EnrichedRows by resolved_filename for one-shot lookup.
    by_fn: dict[str, EnrichedRow] = {}
    for r in rows:
        fn = (r.obs.resolved_filename or "").strip()
        if fn and fn not in by_fn:
            by_fn[fn] = r

    overrides = 0
    rows_seen = 0
    for excel_row in range(2, ws.max_row + 1):
        fn_cell = ws.cell(row=excel_row, column=filename_col).value
        fn = str(fn_cell).strip() if fn_cell else ""
        if not fn or fn not in by_fn:
            continue
        rows_seen += 1
        target = by_fn[fn]
        baseline = (snapshot or {}).get(fn, {})
        for col_idx, hdr in enumerate(headers, start=1):
            if hdr not in _ENRICHED_EDITABLE_COLUMNS:
                continue
            attr = _ENRICHED_EDITABLE_COLUMNS[hdr]
            if attr is None:
                continue
            new_val = ws.cell(row=excel_row, column=col_idx).value
            new_val = "" if new_val is None else str(new_val).strip()
            # When phase 1 recorded a baseline, only treat the cell as
            # edited if the current value differs from what was
            # written. Without a baseline (legacy state files) fall
            # back to the bare attribute comparison.
            if baseline:
                if new_val == baseline.get(hdr, ""):
                    continue
            else:
                old_val = getattr(target, attr, "")
                if new_val == (old_val or ""):
                    continue
            setattr(target, attr, new_val)
            overrides += 1
    return {
        "applied": True,
        "rows_matched": rows_seen,
        "field_overrides": overrides,
    }


def run_once(
    folder: Path,
    prepared_by: str = "Alan Richardson",
    ignore_freeze: bool = False,
    checklist_path: Path | None = None,
    force: bool = False,
    enrich: bool = True,
    risk_assessment_path: Path | None = None,
    merge_groups: list[list[int]] | None = None,
    stop_after: str | None = None,
    from_state: bool = False,
    from_report: bool = False,
) -> dict:
    """Run the pipeline once. Returns the .ssa_run.json payload.

    Idempotency: when ``force`` is False and a prior ``.ssa_run.json``
    is present whose ``inputs_sha256`` matches the current manifest AND
    every recorded output still exists on disk, the run is a no-op and
    the existing payload is returned with ``skipped=True`` set.
    """
    folder = folder.resolve()
    if not folder.is_dir():
        raise NotADirectoryError(folder)

    # Runtime preflight — must happen before any row processing so a
    # missing API key (or other config gap) fails loud rather than
    # producing a structurally-valid but semantically-degraded run.
    _preflight(enrich=enrich)

    freeze = folder / ".ssa_freeze"
    if freeze.exists() and not ignore_freeze:
        raise RuntimeError(
            f"frozen — use --ignore-freeze to overwrite ({freeze})"
        )

    iso, ddmmyyyy, yymmdd, client, sub_id = _parse_folder(folder)
    csv_path = folder / "Evidence_Master.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"Evidence_Master.csv missing in {folder}")

    images = _images_in(folder)
    if not images:
        raise FileNotFoundError(f"no images found in {folder}")

    # --- manifest + idempotency -------------------------------------
    names = _output_names(yymmdd, client, sub_id)
    prior_reports = _qualifying_prior_reports(folder, iso, names["report"])
    manifest = _compute_manifest(csv_path, images, prior_reports)

    prior_record = _existing_run_record(folder)
    if (
        not force
        and prior_record is not None
        and prior_record.get("inputs_sha256") == manifest
        and all((folder / n).exists() for n in prior_record.get("outputs", []))
    ):
        prior_record["skipped"] = True
        log.info("manifest unchanged + outputs present — skipping")
        return prior_record

    # Phase-3 short-circuit: when ``from_report`` is True, skip
    # parse / match / enrichment / vision and rebuild EnrichedRows
    # from the .ssa_state.json sidecar, then apply operator edits
    # from the audit-report docx's per-finding detail tables BACK
    # onto the rows. Re-renders the enriched + staging xlsx files
    # only — leaves the operator-edited docx alone.
    state_path = folder / ".ssa_state.json"
    if from_report:
        if not state_path.exists():
            raise FileNotFoundError(
                f"--from-report requested but .ssa_state.json missing in "
                f"{folder}. Run the pipeline (or phase 1 then phase 2) "
                f"to produce the state sidecar first."
            )
        state = json.loads(state_path.read_text(encoding="utf-8"))
        enriched = _deserialise_rows(state["rows"])
        site_address = state.get("site_address")
        ra_project_name = state.get("ra_project_name", "")
        principal_contractor = state.get("principal_contractor", "")
        finding_render_order = list(state.get("finding_render_order") or [])
        report_path = folder / names["report"]
        enriched_path = folder / names["enriched"]
        staging_path = folder / names["staging"]
        merge_diag = _merge_edits_from_audit_report(
            enriched, report_path, finding_render_order,
        )
        # Rebuild the enriched xlsx + staging from the now-updated rows.
        enriched_diag = build_pims_enriched_xlsx(
            enriched, enriched_path,
            project_name=ra_project_name,
            site_address=site_address or "",
            principal_contractor=principal_contractor,
            audit_date_ddmmyyyy=ddmmyyyy,
        )
        site_for_staging = site_address or ""
        staging_result = build_pims_staging_xlsx_with_size_control(
            enriched, staging_path,
            site_address=site_for_staging,
            audit_date_iso=iso,
            prepared_by=prepared_by,
        )
        staging_diag = {
            "parts":        [p.name for p in staging_result["parts"]],
            "max_edge_px":  staging_result["max_edge_px"],
            "split":        staging_result["split"],
            "split_reason": staging_result["split_reason"],
            "per_part":     staging_result["diagnostics"],
        }
        # Refresh state JSON with the operator-edited rows.
        state["rows"] = _serialise_rows(enriched)
        state_path.write_text(
            json.dumps(state, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        staging_status, blocker = _resolve_staging_status(
            client, site_address,
        )
        outputs = [enriched_path.name, *[p.name for p in staging_result["parts"]]]
        if staging_status == "not_uploadable":
            outputs.append(_write_sentinel(
                folder, "STAGING-NOT-UPLOADABLE.txt",
                f"staging blocker: {blocker}\nfolder: {folder.name}\n",
            ))
        elif staging_status == "schema_valid_no_endpoint":
            outputs.append(_write_sentinel(
                folder, "STAGING-NO-BULK-ENDPOINT.txt",
                f"client: {client}\n",
            ))
        payload = {
            "folder": folder.name,
            "client": client,
            "audit_date": iso,
            "phase": "from-report",
            "from_report": True,
            "merge_diag": merge_diag,
            "staging_status": staging_status,
            "blocker": blocker,
            "client_bulk_endpoint": _BULK_ENDPOINT[client],
            "outputs": outputs,
            "row_count": len(enriched),
            "enriched_diagnostics": enriched_diag,
            "staging_diagnostics": staging_diag,
            "completed_at": datetime.now(timezone.utc)
                .strftime("%Y-%m-%dT%H:%M:%SZ"),
        }
        (folder / ".ssa_run.json").write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return payload

    # Two-phase workflow short-circuit: when ``from_state`` is True,
    # skip parse → match → enrichment → vision and rebuild EnrichedRows
    # from the .ssa_state.json sidecar dropped by phase 1. The
    # operator's edits to the enriched xlsx are picked up via
    # _merge_edits_from_enriched_xlsx so phase 2 sees the human's
    # final values without re-running the LLM.
    state_path = folder / ".ssa_state.json"
    if from_state:
        if not state_path.exists():
            raise FileNotFoundError(
                f"--from-state requested but .ssa_state.json missing in "
                f"{folder}. Run phase 1 first with --enrich-only."
            )
        state = json.loads(state_path.read_text(encoding="utf-8"))
        enriched = _deserialise_rows(state["rows"])
        site_address = state.get("site_address")
        narrative_summary = state.get("narrative_summary", "")
        ra_project_name = state.get("ra_project_name", "")
        principal_contractor = state.get("principal_contractor", "")
        ra_summary = state.get("ra_summary", {})
        llm_diag = state.get("llm_diagnostics", {"enabled": False})
        csv_warnings = []
        match_warnings = []
        prior_reports = []
        # Honour the freshly-set sub_id / output names already computed.
        names_for_phase2 = names
        # Pull operator edits from the enriched xlsx (if present).
        enriched_xlsx_path = folder / names_for_phase2["enriched"]
        merge_diag = _merge_edits_from_enriched_xlsx(
            enriched, enriched_xlsx_path,
            snapshot=state.get("enriched_xlsx_snapshot"),
        )
        # Re-resolve prior-recs from the prior report (cheap).
        prior_recs = []
        prior_audit_date_ddmmyy = ""
        # Find newest qualifying prior report for this folder.
        for p in folder.iterdir():
            if not p.is_file() or p.suffix.lower() != ".docx":
                continue
            mm = re.search(
                r"-(\d{6})-(?:RPD|SDG)(?:-\d{2})?\.docx$", p.name,
            )
            if not mm:
                continue
            try:
                cand_iso = datetime.strptime(mm.group(1), "%y%m%d").date().isoformat()
            except ValueError:
                continue
            if cand_iso < iso and p.name != names_for_phase2["report"]:
                prior_reports.append(p)
        if prior_reports:
            newest_prior = sorted(prior_reports)[-1]
            prior_recs = parse_prior_report_recommendations(newest_prior)
            mm = re.search(
                r"-(\d{6})-(?:RPD|SDG)(?:-\d{2})?\.docx$", newest_prior.name,
            )
            if mm:
                yymmdd = mm.group(1)
                prior_audit_date_ddmmyy = (
                    f"{yymmdd[4:6]}/{yymmdd[2:4]}/{yymmdd[0:2]}"
                )
        # Recompute paths for the from_state branch and skip the
        # rest of the parse/enrich block by jumping straight to the
        # build step. We do this by setting a flag and falling through.
        enriched_path = folder / names_for_phase2["enriched"]
        report_path = folder / names_for_phase2["report"]
        staging_path = folder / names_for_phase2["staging"]
        site_for_docx = site_address or "[Site address - to be confirmed]"
        site_for_staging = site_address or ""
        # Build only the report + staging (skip enriched xlsx so the
        # operator's edits stay intact).
        report_diag = build_ssa_report_docx(
            enriched,
            site_address=site_for_docx,
            audit_date_ddmmyyyy=ddmmyyyy,
            narrative_summary=narrative_summary,
            output_path=report_path,
            prepared_by=prepared_by,
            prior_recs=prior_recs,
            project_name=ra_project_name,
            prior_audit_date_ddmmyy=prior_audit_date_ddmmyy,
            risk_assessment=None,
            merge_groups=merge_groups,
        )
        report_diag["prior_recs_count"] = len(prior_recs)
        staging_result = build_pims_staging_xlsx_with_size_control(
            enriched,
            staging_path,
            site_address=site_for_staging,
            audit_date_iso=iso,
            prepared_by=prepared_by,
        )
        staging_diag = {
            "parts":         [p.name for p in staging_result["parts"]],
            "max_edge_px":   staging_result["max_edge_px"],
            "split":         staging_result["split"],
            "split_reason":  staging_result["split_reason"],
            "per_part":      staging_result["diagnostics"],
        }
        staging_status, blocker = _resolve_staging_status(client, site_address)
        outputs: list[str] = [
            enriched_path.name, report_path.name,
            *[p.name for p in staging_result["parts"]],
        ]
        if staging_status == "not_uploadable":
            outputs.append(_write_sentinel(
                folder, "STAGING-NOT-UPLOADABLE.txt",
                f"staging blocker: {blocker}\nfolder: {folder.name}\n",
            ))
        elif staging_status == "schema_valid_no_endpoint":
            outputs.append(_write_sentinel(
                folder, "STAGING-NO-BULK-ENDPOINT.txt",
                f"client: {client}\n",
            ))
        payload = {
            "folder": folder.name,
            "client": client,
            "audit_date": iso,
            "inputs_sha256": manifest,
            "prior_reports_used": [p.name for p in prior_reports],
            "skipped": False,
            "from_state": True,
            "merge_diag": merge_diag,
            "staging_status": staging_status,
            "blocker": blocker,
            "client_bulk_endpoint": _BULK_ENDPOINT[client],
            "outputs": outputs,
            "row_count": len(enriched),
            "csv_warnings": [],
            "match_warnings": [],
            "review_reasons_per_row": [],
            "enriched_diagnostics": {
                "phase2_skipped": "operator-edited xlsx kept intact",
            },
            "report_diagnostics": report_diag,
            "staging_diagnostics": staging_diag,
            "llm_diagnostics": llm_diag,
            "completed_at": datetime.now(timezone.utc)
                .strftime("%Y-%m-%dT%H:%M:%SZ"),
        }
        (folder / ".ssa_run.json").write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return payload

    # --- parse + match + address ------------------------------------
    rows, csv_warnings = parse_evidence_csv(csv_path)
    match_warnings = match_photos(rows, images)

    # Item 11: split composite "(1) X (2) Y" notes into atomic
    # observations BEFORE photo-match metadata is consumed downstream.
    # Each split shares the same csv_row, photo and timestamp; only
    # the observation_text differs.
    rows = split_multi_issue_observations(rows)

    # Gap-8: Vision is the canonical classifier. The legacy keyword
    # matcher in ChecklistLookup.match_observation produced 5/21 hits
    # with one outright misroute on real audit data — it's only kept
    # for the offline (--no-enrich) path AND only when the operator
    # explicitly passes --checklist. The default (vision) path skips
    # the xlsx load entirely so a missing audit_checklist.xlsx never
    # silently degrades the run.
    checklist = None
    if not enrich and checklist_path is not None:
        if checklist_path.exists():
            checklist = ChecklistLookup.from_xlsx(checklist_path)
        else:
            log.warning(
                "--checklist %s does not exist; offline run continues "
                "with no keyword fallback", checklist_path,
            )

    site_address = extract_site_address(rows)

    # When vision is on, enrich_observations builds shells only — the
    # vision pass downstream does the real classification. When vision
    # is off and a checklist was loaded, enable the keyword auto-match
    # path so the run produces something better than all-Unmatched.
    enriched = enrich_observations(
        rows,
        checklist=checklist,
        auto_match=(checklist is not None and not enrich),
    )

    # When site_address is unresolved, every staging row must
    # needs_review=TRUE (Field Defaults). enrich_observations already
    # sets needs_review for Unmatched rows; flagging the obs surfaces
    # the address-unresolved reason in the warning trail.
    if site_address is None:
        for r in enriched:
            r.obs.flag("site_address_unresolved")

    # --- build deliverables -----------------------------------------
    enriched_path = folder / names["enriched"]
    report_path = folder / names["report"]
    staging_path = folder / names["staging"]

    site_for_docx = site_address or "[Site address - to be confirmed]"
    site_for_staging = site_address or ""

    # Project Risk Assessment context (optional). Auto-discovers any
    # ``*Risk_Assessment*.docx`` in the audit folder when no explicit
    # path is supplied. Empty string when no RA is available — the
    # vision call falls back to generic Australian-WHS classification.
    from pims.services.ssa_ra_parser import (
        autodiscover_in_folder, compact_context_block, parse_risk_assessment,
    )
    ra_path = risk_assessment_path
    if ra_path is None:
        ra_path = autodiscover_in_folder(folder)
    ra_context = ""
    ra_project_name = ""
    ra_obj = None
    ra_summary: dict[str, object] = {"path": None, "phases": 0, "activities": 0,
                                     "hold_points": 0}
    if ra_path is not None:
        ra = parse_risk_assessment(ra_path)
        ra_obj = ra
        ra_context = compact_context_block(ra)
        # Strip the "— N Industrial Warehouse Units" suffix that some
        # RA project names carry; the cover line wants just the venue.
        if ra.project_name:
            for sep in (" — ", " – ", " - "):
                if sep in ra.project_name:
                    ra_project_name = ra.project_name.split(sep, 1)[0].strip()
                    break
            else:
                ra_project_name = ra.project_name.strip()
        ra_summary = {
            "path": ra_path.name,
            "project": ra.project_name,
            "phases": len({a.phase for a in ra.activities}),
            "activities": len(ra.activities),
            "hold_points": len(ra.hold_points),
        }
        log.info(
            "RA loaded: %s — %d activities, %d hold points",
            ra_path.name, len(ra.activities), len(ra.hold_points),
        )

    # Vision enrichment — per-row classification (status, CCVS code,
    # finding text, legal_ref, recommendation, monitoring_note) plus
    # the Executive Summary paragraph. Default-on. No-op when
    # ``--no-enrich`` was passed or ANTHROPIC_API_KEY is unset.
    narrative_summary, llm_diag = _apply_vision_enrichment(
        enriched,
        site_address=site_address,
        audit_date_iso=iso,
        enable=enrich,
        ra_context=ra_context,
    )
    llm_diag["ra"] = ra_summary

    # Items 9 + 14: apply "SDG Project Risk Assessment code: <CODE>"
    # labelling with first-use shorthand expansion to every row's
    # text fields. Runs once after vision enrichment so the
    # downstream enriched xlsx / docx / staging xlsx all carry the
    # same labelled output.
    apply_ra_labels_to_rows(enriched, ra=ra_obj)
    # Item 9: label RA codes inside the executive summary too.
    from pims.services.ssa_pipeline import apply_ra_code_labels
    if narrative_summary:
        narrative_summary = apply_ra_code_labels(narrative_summary, ra=ra_obj)

    # Parse carry-forward recommendations from the newest qualifying
    # prior report so the SSA report's "Status of Previous
    # Recommendations" table actually carries content.
    prior_recs: list[dict] = []
    prior_audit_date_ddmmyy = ""
    if prior_reports:
        newest_prior = prior_reports[-1]
        prior_recs = parse_prior_report_recommendations(newest_prior)
        # Extract YYMMDD date from filename, format as DD/MM/YY for the
        # canonical "Status (DD/MM/YY)" header in the prior-recs table.
        m = re.search(
            r"-(\d{6})-(?:RPD|SDG)(?:-\d{2})?\.docx$", newest_prior.name,
        )
        if m:
            yymmdd = m.group(1)
            prior_audit_date_ddmmyy = (
                f"{yymmdd[4:6]}/{yymmdd[2:4]}/{yymmdd[0:2]}"
            )

    # Pull principal contractor + project metadata from the parsed RA
    # so the Enriched workbook's Summary sheet matches the canonical
    # sample's title block / metadata rows.
    principal_contractor = ""
    if ra_path is not None:
        # Re-parse for the metadata only (compact_context_block already
        # consumed the parsed object once). Cheap; one xlsx-style read.
        try:
            ra_meta = parse_risk_assessment(ra_path)
            principal_contractor = ra_meta.principal_contractor
        except Exception:
            log.warning("RA principal-contractor lookup failed", exc_info=True)

    enriched_diag = build_pims_enriched_xlsx(
        enriched, enriched_path,
        project_name=ra_project_name,
        site_address=site_address or "",
        principal_contractor=principal_contractor,
        audit_date_ddmmyyyy=ddmmyyyy,
    )

    # render_order is filled by build_ssa_report_docx to map detail
    # table position → csv_idx. Phase-3 (--from-report) pairs docx
    # edits back to EnrichedRows using this list.
    finding_render_order: list[int] = []

    # Persist the full row state for the two-phase workflow. Phase 2
    # (--from-state) reads this back, optionally merges operator
    # edits from the enriched xlsx, then renders the report + staging.
    # Snapshot the enriched xlsx's editable cells AS WRITTEN so
    # phase 2 can distinguish operator edits from rendered-fallback
    # values that look different from the source attribute (e.g. the
    # Action Description cell shows ``recommendation`` when
    # ``action_description`` is empty).
    enriched_snapshot = _snapshot_enriched_xlsx(enriched_path)

    state_payload = {
        "folder": folder.name,
        "client": client,
        "audit_date": iso,
        "audit_date_ddmmyyyy": ddmmyyyy,
        "site_address": site_address,
        "ra_project_name": ra_project_name,
        "principal_contractor": principal_contractor,
        "ra_summary": ra_summary,
        "narrative_summary": narrative_summary,
        "llm_diagnostics": llm_diag,
        "rows": _serialise_rows(enriched),
        "enriched_xlsx_snapshot": enriched_snapshot,
        "finding_render_order": list(finding_render_order),
    }
    state_path.write_text(
        json.dumps(state_payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    # Phase-1 short-circuit: when ``stop_after="enrich"`` the pipeline
    # writes the enriched xlsx + state JSON and exits BEFORE the report
    # / staging are produced. The operator reviews + edits the xlsx,
    # then runs phase 2 with --from-state.
    if stop_after == "enrich":
        sentinel_outputs = [enriched_path.name, ".ssa_state.json"]
        ph1_payload = {
            "folder": folder.name,
            "client": client,
            "audit_date": iso,
            "phase": "enrich-only",
            "outputs": sentinel_outputs,
            "row_count": len(enriched),
            "llm_diagnostics": llm_diag,
            "site_address": site_address,
            "site_address_unresolved": site_address is None,
            "ra": ra_summary,
            "completed_at": datetime.now(timezone.utc)
                .strftime("%Y-%m-%dT%H:%M:%SZ"),
            "next_step": (
                "review and edit the Enriched Register sheet in "
                f"{enriched_path.name}, then re-run with "
                "--from-state to produce the report + staging files"
            ),
        }
        (folder / ".ssa_run.json").write_text(
            json.dumps(ph1_payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return ph1_payload
    report_diag = build_ssa_report_docx(
        enriched,
        site_address=site_for_docx,
        audit_date_ddmmyyyy=ddmmyyyy,
        narrative_summary=narrative_summary,
        output_path=report_path,
        prepared_by=prepared_by,
        prior_recs=prior_recs,
        project_name=ra_project_name,
        prior_audit_date_ddmmyy=prior_audit_date_ddmmyy,
        risk_assessment=ra_obj,
        merge_groups=merge_groups,
        render_order_out=finding_render_order,
    )
    report_diag["prior_recs_count"] = len(prior_recs)

    # Update the state JSON with the freshly-captured
    # finding_render_order so phase-3 (--from-report) can pair docx
    # detail-table edits back to the correct EnrichedRow.
    if state_path.exists():
        try:
            existing_state = json.loads(
                state_path.read_text(encoding="utf-8"),
            )
            existing_state["finding_render_order"] = list(
                finding_render_order,
            )
            existing_state["rows"] = _serialise_rows(enriched)
            state_path.write_text(
                json.dumps(existing_state, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            log.warning("failed to refresh state JSON", exc_info=True)

    staging_result = build_pims_staging_xlsx_with_size_control(
        enriched,
        staging_path,
        site_address=site_for_staging,
        audit_date_iso=iso,
        prepared_by=prepared_by,
    )
    staging_diag = {
        "parts":         [p.name for p in staging_result["parts"]],
        "max_edge_px":   staging_result["max_edge_px"],
        "split":         staging_result["split"],
        "split_reason":  staging_result["split_reason"],
        "per_part":      staging_result["diagnostics"],
    }

    # --- staging tri-state ------------------------------------------
    staging_status, blocker = _resolve_staging_status(client, site_address)
    staging_part_names = [p.name for p in staging_result["parts"]]
    outputs: list[str] = [
        enriched_path.name,
        report_path.name,
        *staging_part_names,
    ]

    if staging_status == "not_uploadable":
        outputs.append(_write_sentinel(
            folder, "STAGING-NOT-UPLOADABLE.txt",
            f"staging blocker: {blocker}\n"
            f"folder: {folder.name}\n"
            f"remediation: see .claude/plans/workflow-1 — for "
            f"site_address_unresolved, add an address-shaped sentence to "
            f"a row in Evidence_Master.csv and rerun.\n",
        ))
    elif staging_status == "schema_valid_no_endpoint":
        outputs.append(_write_sentinel(
            folder, "STAGING-NO-BULK-ENDPOINT.txt",
            f"client: {client}\n"
            f"the staging xlsx is schema-valid and forward-compatible.\n"
            f"no bulk-upload endpoint exists for {client} today; post "
            f"observations one-at-a-time via the single-observation API.\n",
        ))

    # --- run record --------------------------------------------------
    payload = {
        "folder": folder.name,
        "client": client,
        "audit_date": iso,
        "inputs_sha256": manifest,
        "prior_reports_used": [p.name for p in prior_reports],
        "skipped": False,
        "prepared_by": prepared_by,
        "site_address": site_address,
        "site_address_unresolved": site_address is None,
        "staging_status": staging_status,
        "blocker": blocker,
        "client_bulk_endpoint": _BULK_ENDPOINT[client],
        "outputs": outputs,
        "row_count": len(enriched),
        "csv_warnings": [w.to_dict() for w in csv_warnings],
        "match_warnings": [w.to_dict() for w in match_warnings],
        "review_reasons_per_row": [
            {"csv_row": r.obs.csv_row, "reasons": list(r.obs.review_reasons)}
            for r in enriched if r.obs.review_reasons
        ],
        "enriched_diagnostics": enriched_diag,
        "report_diagnostics": report_diag,
        "staging_diagnostics": staging_diag,
        "llm_diagnostics": llm_diag,
        "completed_at": datetime.now(timezone.utc)
            .strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    (folder / ".ssa_run.json").write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return payload


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="run_ssa_pipeline")
    ap.add_argument("folder", type=Path, help="audit evidence folder")
    ap.add_argument("--prepared-by", default="Alan Richardson")
    ap.add_argument("--ignore-freeze", action="store_true")
    ap.add_argument(
        "--force", action="store_true",
        help="ignore the manifest-based idempotency skip",
    )
    ap.add_argument("--checklist", type=Path, default=None)
    ap.add_argument(
        "--risk-assessment", type=Path, default=None,
        help="path to project Risk Assessment .docx; "
             "auto-discovered from audit folder when omitted",
    )
    ap.add_argument(
        "--no-enrich", action="store_true",
        help="skip the LLM finding-rewrite + narrative-summary pass",
    )
    ap.add_argument(
        "--enrich-only", action="store_true",
        help=(
            "phase 1 of the two-phase workflow — write the enriched "
            "xlsx + .ssa_state.json sidecar and EXIT before the "
            "report and staging files are produced. Use this when "
            "you want to review / edit the LLM's findings in Excel "
            "before they get baked into the docx + staging xlsx. "
            "Re-run with --from-state to complete the run."
        ),
    )
    ap.add_argument(
        "--from-state", action="store_true",
        help=(
            "phase 2 of the two-phase workflow — skip parse / match "
            "/ vision and rebuild EnrichedRows from the .ssa_state."
            "json sidecar, merging the operator's edits from the "
            "Enriched Register sheet. Produces the report + staging "
            "files; leaves the (already operator-edited) enriched "
            "xlsx untouched."
        ),
    )
    ap.add_argument(
        "--from-report", action="store_true",
        help=(
            "phase 3 of the workflow — read operator edits from the "
            "audit report docx (per-finding detail tables) and flow "
            "them BACK to the enriched + staging xlsx. Updates the "
            "Location / Observation / Regulatory Basis / Hierarchy "
            "of Control / Recommendation / Timeframe fields. Leaves "
            "the operator-edited docx alone (it's the source of "
            "truth)."
        ),
    )
    ap.add_argument(
        "--merge", default="",
        help=(
            "merge directives by displayed finding number, e.g. "
            "\"1,3\" merges findings #1 and #3 into one consolidated "
            "entry; \"1,3;5,7\" applies two merges. Indices are "
            "1-based and refer to the post-significance order shown "
            "in the docx Findings index table."
        ),
    )
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s | %(message)s",
    )

    try:
        phase_flags = (args.enrich_only, args.from_state, args.from_report)
        if sum(bool(f) for f in phase_flags) > 1:
            print(
                "error: --enrich-only / --from-state / --from-report "
                "are mutually exclusive (they select one phase of the "
                "review workflow).",
                file=sys.stderr,
            )
            return 1
        payload = run_once(
            args.folder,
            prepared_by=args.prepared_by,
            ignore_freeze=args.ignore_freeze,
            checklist_path=args.checklist,
            force=args.force,
            enrich=not args.no_enrich,
            risk_assessment_path=args.risk_assessment,
            merge_groups=parse_merge_argument(args.merge),
            stop_after="enrich" if args.enrich_only else None,
            from_state=args.from_state,
            from_report=args.from_report,
        )
    except PreflightError as e:
        # Preflight blocked the run before any rows processed; rc=3
        # distinguishes a config gap from a frozen folder (rc=2) and
        # an input/argument error (rc=1).
        print(f"preflight: {e}", file=sys.stderr)
        return 3
    except RuntimeError as e:
        # Frozen folder — the documented exit signal for the manual CLI
        # is non-zero with a clear message.
        print(f"error: {e}", file=sys.stderr)
        return 2
    except (FileNotFoundError, NotADirectoryError, ValueError) as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    if payload.get("skipped"):
        print("skipped: manifest unchanged + all outputs present")
    if payload.get("phase") == "enrich-only":
        print("phase 1 (enrich-only) complete")
        print(f"row_count: {payload.get('row_count')}")
        print(f"next step: {payload.get('next_step', '')}")
        print(f"outputs ({len(payload['outputs'])} files):")
        for name in payload["outputs"]:
            print(f"  - {name}")
        return 0
    print(f"staging_status: {payload['staging_status']}")
    if payload.get("blocker"):
        print(f"blocker:        {payload['blocker']}")
    print(f"outputs:        {len(payload['outputs'])} files in {args.folder}")
    for name in payload["outputs"]:
        print(f"  - {name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
