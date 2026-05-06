"""Interactive review + edit CLI for the SSA enrichment phase.

Sits between phase 1 (``--enrich-only``) and phase 2 (``--from-state``)
of the SSA workflow. The operator reviews the enriched.xlsx +
state.json output, applies edits via direct field commands or
plain-English directives, then triggers phase 2 to render the
report and staging xlsx.

Subcommands:

  list                       show every finding with its number,
                             status, ccvs and short title
  show <n>                   dump every editable field of finding <n>
  set <n> <field> <value>    direct field edit:
                             set 3 status NCR
                             set 7 ccvs WAH-H6
                             set 4 finding "Edge protection missing
                             on level 2"
  find <text>                grep across observation / finding text
  nl "<directive>"           plain-English directive parsed via a
                             small Anthropic call:
                             nl "change row 4 status to NCR and add
                             HRCW H15 to the legal ref"
  render                     run phase 2 (--from-state) — produces
                             docx + staging xlsx with current edits
  repl                       interactive shell (default when no
                             subcommand supplied)

State persistence: every edit immediately rewrites the
``Enriched Register`` cell of the enriched xlsx AND the
``.ssa_state.json`` row, so phase 2 picks up the change without any
extra step. The original photo / timestamp / csv-row metadata is
never touched.

iPhone usage: connect to your Mac via SSH (Termius / Blink), open
this CLI's REPL with ``ssa_review repl <folder>``, dictate
directives via iOS Dictation. The REPL prints confirmations after
each edit so you can verify by ear.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shlex
import subprocess
import sys
from pathlib import Path

import openpyxl

# Editable column -> state-row attribute. Mirrors
# _ENRICHED_EDITABLE_COLUMNS in run_ssa_pipeline.py but as a 1:1
# mapping (no Nones). Keys are case-insensitive lowercase header text.
EDITABLE_FIELDS: dict[str, str] = {
    "observation": "finding",
    "conformance status": "conformance_status",
    "status": "conformance_status",         # alias
    "ccvs code": "ccvs_code",
    "ccvs": "ccvs_code",                     # alias
    "ccvs category": "ccvs_category",
    "category": "ccvs_category",            # alias
    "action description": "action_description",
    "action": "action_description",          # alias
    "monitoring note": "monitoring_note",
    "monitoring": "monitoring_note",         # alias
    "responsible": "responsible",            # blank-by-default cell
    "due": "due",                            # blank-by-default cell
    "finding": "finding",                    # treat raw "finding" the
                                             # same as the Observation
                                             # column for ergonomics
    "title": "finding_title",                # phase-3 title edit
    "location": "location",
    "recommendation": "recommendation",
    "hierarchy": "hierarchy_of_control",
    "legal ref": "legal_ref",
    "legal": "legal_ref",                    # alias
    "timeframe": "timeframe",
}

ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_VERSION = "2023-06-01"
NL_MODEL = "claude-opus-4-7"


# ---------------------------------------------------------------------------
# State + xlsx loaders
# ---------------------------------------------------------------------------

def _load_state(folder: Path) -> dict:
    state_path = folder / ".ssa_state.json"
    if not state_path.exists():
        raise FileNotFoundError(
            f"{state_path} not found. Run phase 1 first: "
            f"`python -m pims.scripts.run_ssa_pipeline \"{folder}\" "
            f"--force --enrich-only`."
        )
    return json.loads(state_path.read_text(encoding="utf-8"))


def _save_state(folder: Path, state: dict) -> None:
    (folder / ".ssa_state.json").write_text(
        json.dumps(state, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _enriched_xlsx_path(folder: Path, state: dict) -> Path:
    """Output names depend on folder name + sub-id; reuse the same
    derivation the orchestrator does."""
    m = re.match(
        r"^(\d{4})-(\d{2})-(\d{2})-(RPD|SDG)(?:-(\d{2}))?$", folder.name,
    )
    if not m:
        raise ValueError(f"folder {folder.name!r} doesn't match SSA naming")
    yyyy, mm, dd, client, sub_id = m.groups()
    yymmdd = f"{yyyy[2:]}{mm}{dd}"
    suffix = f"-{client}-{sub_id}" if sub_id else f"-{client}"
    return folder / f"PIMS-Enriched-{yymmdd}{suffix}.xlsx"


# ---------------------------------------------------------------------------
# Find / show / list
# ---------------------------------------------------------------------------

def _row_number_to_state_idx(state: dict, n: int) -> int:
    """Turn a 1-based finding number into a 0-based state index. The
    enriched xlsx writes data rows from row 2 in CSV order, so finding
    #N lives at state["rows"][N-1].
    """
    rows = state.get("rows") or []
    if n < 1 or n > len(rows):
        raise IndexError(
            f"finding #{n} out of range — there are {len(rows)} rows"
        )
    return n - 1


def _short_title(row: dict) -> str:
    return (
        row.get("finding_title")
        or row.get("ccvs_category")
        or "Finding"
    )


def cmd_list(state: dict) -> str:
    rows = state.get("rows") or []
    if not rows:
        return "(no rows)"
    out = []
    for i, r in enumerate(rows, start=1):
        out.append(
            f"  #{i:<3} {r.get('conformance_status', '?'):<13} "
            f"{r.get('ccvs_code', ''):<10} {_short_title(r)[:70]}"
        )
    return "\n".join(out)


def cmd_show(state: dict, n: int) -> str:
    idx = _row_number_to_state_idx(state, n)
    r = state["rows"][idx]
    fields = [
        ("status", r.get("conformance_status")),
        ("ccvs", r.get("ccvs_code")),
        ("category", r.get("ccvs_category")),
        ("title", r.get("finding_title")),
        ("location", r.get("location")),
        ("finding", r.get("finding")),
        ("legal_ref", r.get("legal_ref")),
        ("hierarchy", r.get("hierarchy_of_control")),
        ("recommendation", r.get("recommendation")),
        ("timeframe", r.get("timeframe")),
        ("monitoring_note", r.get("monitoring_note")),
        ("action_description", r.get("action_description")),
    ]
    head = (
        f"finding #{n} (csv_row={r.get('obs', {}).get('csv_row')}, "
        f"photo={r.get('obs', {}).get('resolved_filename') or '—'})"
    )
    body = "\n".join(
        f"  {label:<18} {value or ''}" for label, value in fields
    )
    return f"{head}\n{body}"


def cmd_find(state: dict, query: str) -> str:
    rows = state.get("rows") or []
    q = query.lower()
    hits = []
    for i, r in enumerate(rows, start=1):
        blob = " ".join((
            r.get("finding") or "",
            r.get("observation_text_clean") or "",
            r.get("finding_title") or "",
            r.get("ccvs_code") or "",
            r.get("ccvs_category") or "",
        )).lower()
        if q in blob:
            hits.append(
                f"  #{i:<3} {_short_title(r)[:70]}"
            )
    return "\n".join(hits) if hits else f"no rows match {query!r}"


# ---------------------------------------------------------------------------
# set / direct edit
# ---------------------------------------------------------------------------

def _resolve_field(field_arg: str) -> str:
    """Map a user-supplied field name (case-insensitive, with aliases)
    onto the canonical row attribute. Raises KeyError on unknowns."""
    key = field_arg.strip().lower()
    if key in EDITABLE_FIELDS:
        return EDITABLE_FIELDS[key]
    raise KeyError(
        f"unknown field {field_arg!r}. "
        f"Try one of: {', '.join(sorted(set(EDITABLE_FIELDS)))}"
    )


def _apply_edit(
    state: dict, n: int, field_attr: str, new_value: str,
) -> tuple[str, str]:
    """Mutate state['rows'][n-1][field_attr] = new_value. Returns
    (old_value, new_value) for confirmation."""
    idx = _row_number_to_state_idx(state, n)
    row = state["rows"][idx]
    old = row.get(field_attr, "")
    row[field_attr] = new_value
    return (str(old or ""), str(new_value))


def _write_enriched_cell(
    folder: Path, state: dict, n: int, field_attr: str, new_value: str,
) -> bool:
    """Update the matching cell in the enriched xlsx so it stays in
    sync with the state JSON. Returns True when a cell was written.
    """
    xlsx_path = _enriched_xlsx_path(folder, state)
    if not xlsx_path.exists():
        return False
    # Reverse-map the attribute back to the xlsx header text.
    header_to_attr = {
        "observation": "finding",
        "conformance status": "conformance_status",
        "ccvs code": "ccvs_code",
        "ccvs category": "ccvs_category",
        "action description": "action_description",
        "monitoring note": "monitoring_note",
        "responsible": "responsible",
        "due": "due",
    }
    target_header = next(
        (h for h, a in header_to_attr.items() if a == field_attr),
        None,
    )
    if target_header is None:
        return False
    wb = openpyxl.load_workbook(xlsx_path)
    if "Enriched Register" not in wb.sheetnames:
        return False
    ws = wb["Enriched Register"]
    headers = [
        ("" if c.value is None else str(c.value).strip().lower())
        for c in ws[1]
    ]
    if target_header not in headers:
        return False
    col_idx = headers.index(target_header) + 1
    excel_row = n + 1  # row 1 = header, data starts row 2
    ws.cell(row=excel_row, column=col_idx, value=new_value)
    wb.save(xlsx_path)
    return True


def cmd_set(folder: Path, state: dict, n: int, field: str, value: str) -> str:
    field_attr = _resolve_field(field)
    old, new = _apply_edit(state, n, field_attr, value)
    wrote = _write_enriched_cell(folder, state, n, field_attr, value)
    _save_state(folder, state)
    return (
        f"#{n} {field_attr}: {old!r} -> {new!r}"
        f" {'(xlsx updated)' if wrote else '(state only — not in xlsx schema)'}"
    )


# ---------------------------------------------------------------------------
# Natural-language directive parser
# ---------------------------------------------------------------------------

_NL_SYSTEM = (
    "You translate one auditor's voice directive into a list of edits "
    "against the SSA Enriched Register. Output JSON only, no prose. "
    "Schema: {\"edits\": [{\"row\": <int>, \"field\": <string>, \"value\": "
    "<string>}, ...]}.\n\n"
    "row is the 1-based finding number the auditor refers to (e.g. "
    "\"row 4\", \"finding 7\", \"#3\"). When the auditor describes the "
    "row by content (\"the fire extinguisher one\") the caller is "
    "responsible for resolving — output the row number they "
    "explicitly said, otherwise leave row=0 and field=\"\".\n\n"
    "field must be one of: observation, conformance status, ccvs code, "
    "ccvs category, action description, monitoring note, responsible, "
    "due, title, location, recommendation, hierarchy, legal ref, "
    "timeframe, finding. Aliases like \"status\" -> \"conformance "
    "status\", \"ccvs\" -> \"ccvs code\", \"category\" -> \"ccvs "
    "category\" are fine; the caller normalises.\n\n"
    "value is the literal new cell content. For status, use exactly "
    "one of: Compliant, Conditional, NCR, Info, Unmatched. For ccvs "
    "code, the canonical <STREAM>-<TIER> form (e.g. WAH-H6, MOB-H9, "
    "SYS-L1).\n\n"
    "When the directive contains multiple edits, return one entry "
    "per edit. Examples:\n"
    "  IN: \"change row 4 status to NCR and add HRCW H15 to the legal ref\"\n"
    "  OUT: {\"edits\": [{\"row\": 4, \"field\": \"conformance status\", "
    "\"value\": \"NCR\"}, {\"row\": 4, \"field\": \"legal ref\", "
    "\"value\": \"HRCW H15\"}]}\n"
    "  IN: \"set finding 7 ccvs to MOB-H9\"\n"
    "  OUT: {\"edits\": [{\"row\": 7, \"field\": \"ccvs code\", "
    "\"value\": \"MOB-H9\"}]}"
)


def _nl_call(directive: str) -> dict:
    """One Anthropic call to translate ``directive`` to structured
    edits. Raises on missing key / network failure."""
    import httpx
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError(
            "ANTHROPIC_API_KEY missing — load .env or set it before "
            "using `nl`."
        )
    body = {
        "model": NL_MODEL,
        "max_tokens": 600,
        "system": _NL_SYSTEM,
        "messages": [{"role": "user", "content": directive}],
    }
    with httpx.Client(timeout=30) as client:
        resp = client.post(
            ANTHROPIC_URL,
            headers={
                "x-api-key": api_key,
                "anthropic-version": ANTHROPIC_VERSION,
                "content-type": "application/json",
            },
            json=body,
        )
        resp.raise_for_status()
        text = resp.json()["content"][0]["text"].strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.rsplit("```", 1)[0].strip()
    return json.loads(text)


def cmd_nl(folder: Path, state: dict, directive: str) -> str:
    payload = _nl_call(directive)
    edits = payload.get("edits") or []
    if not edits:
        return "no edits parsed from directive"
    lines = []
    applied = 0
    for e in edits:
        n = int(e.get("row") or 0)
        field = str(e.get("field") or "").strip()
        value = str(e.get("value") or "")
        if n < 1 or not field:
            lines.append(f"  ! skipped (incomplete): {e!r}")
            continue
        try:
            field_attr = _resolve_field(field)
            old, new = _apply_edit(state, n, field_attr, value)
            wrote = _write_enriched_cell(folder, state, n, field_attr, value)
            applied += 1
            lines.append(
                f"  #{n} {field_attr}: {old[:40]!r} -> {new[:60]!r}"
                f" {'(xlsx)' if wrote else '(state)'}"
            )
        except (KeyError, IndexError) as exc:
            lines.append(f"  ! {exc}")
    if applied:
        _save_state(folder, state)
    head = f"applied {applied}/{len(edits)} edit(s):"
    return head + "\n" + "\n".join(lines)


# ---------------------------------------------------------------------------
# Render — call phase 2
# ---------------------------------------------------------------------------

def cmd_render(folder: Path) -> str:
    """Run phase 2 (`--from-state`) on the same folder so the docx +
    staging files reflect the current edits."""
    cmd = [
        sys.executable, "-m", "pims.scripts.run_ssa_pipeline",
        str(folder), "--force", "--from-state",
    ]
    rc = subprocess.run(cmd, capture_output=True, text=True)
    if rc.returncode != 0:
        return f"render failed (rc={rc.returncode}): {rc.stderr.strip()[:300]}"
    return "render OK\n" + rc.stdout


# ---------------------------------------------------------------------------
# REPL
# ---------------------------------------------------------------------------

_HELP = """
SSA Review REPL — commands:
  list                    show every finding (number, status, ccvs, title)
  show <n>                dump every editable field of finding <n>
  set <n> <field> <value> direct edit (e.g. set 3 status NCR)
  find <text>             grep across observation / finding text
  nl "<directive>"        natural-language edit (Anthropic-parsed)
  render                  run phase 2 to produce docx + staging
  help                    this message
  quit / exit             leave the REPL
"""


def repl(folder: Path) -> int:
    state = _load_state(folder)
    print(f"loaded {len(state.get('rows', []))} findings from {folder.name}")
    print("type 'help' for commands; 'quit' to exit")
    while True:
        try:
            line = input("ssa> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            return 0
        if not line:
            continue
        if line in ("quit", "exit", "q"):
            return 0
        if line == "help":
            print(_HELP)
            continue
        if line == "list":
            print(cmd_list(state))
            continue
        if line == "render":
            print(cmd_render(folder))
            # Reload state in case render mutated it.
            state = _load_state(folder)
            continue
        # Token-based dispatch for: show / set / find / nl
        try:
            tokens = shlex.split(line)
        except ValueError as exc:
            print(f"  ! parse error: {exc}")
            continue
        if not tokens:
            continue
        cmd = tokens[0].lower()
        try:
            if cmd == "show" and len(tokens) >= 2:
                print(cmd_show(state, int(tokens[1])))
            elif cmd == "find" and len(tokens) >= 2:
                print(cmd_find(state, " ".join(tokens[1:])))
            elif cmd == "set" and len(tokens) >= 4:
                n = int(tokens[1])
                field = tokens[2]
                value = " ".join(tokens[3:])
                print(cmd_set(folder, state, n, field, value))
            elif cmd == "nl" and len(tokens) >= 2:
                print(cmd_nl(folder, state, " ".join(tokens[1:])))
            else:
                print(f"  ! unknown command {cmd!r}; type 'help'")
        except (KeyError, IndexError, ValueError) as exc:
            print(f"  ! {exc}")
        except Exception as exc:
            print(f"  ! {type(exc).__name__}: {exc}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(prog="ssa_review")
    ap.add_argument("folder", type=Path, help="audit evidence folder")
    sub = ap.add_subparsers(dest="cmd")

    sub.add_parser("list", help="show every finding")

    p_show = sub.add_parser("show", help="dump finding details")
    p_show.add_argument("n", type=int)

    p_set = sub.add_parser("set", help="direct field edit")
    p_set.add_argument("n", type=int)
    p_set.add_argument("field")
    p_set.add_argument("value", nargs="+")

    p_find = sub.add_parser("find", help="grep findings")
    p_find.add_argument("query", nargs="+")

    p_nl = sub.add_parser("nl", help="natural-language directive")
    p_nl.add_argument("directive", nargs="+")

    sub.add_parser("render", help="run phase 2 (docx + staging)")
    sub.add_parser("repl", help="interactive shell (default)")

    args = ap.parse_args(argv)

    if not args.folder.is_dir():
        print(f"error: {args.folder} not a directory")
        return 1

    cmd = args.cmd or "repl"
    try:
        if cmd == "repl":
            return repl(args.folder)
        state = _load_state(args.folder)
        if cmd == "list":
            print(cmd_list(state))
        elif cmd == "show":
            print(cmd_show(state, args.n))
        elif cmd == "set":
            print(cmd_set(
                args.folder, state, args.n, args.field, " ".join(args.value),
            ))
        elif cmd == "find":
            print(cmd_find(state, " ".join(args.query)))
        elif cmd == "nl":
            print(cmd_nl(args.folder, state, " ".join(args.directive)))
        elif cmd == "render":
            print(cmd_render(args.folder))
        return 0
    except FileNotFoundError as exc:
        print(f"error: {exc}")
        return 1
    except (KeyError, IndexError, ValueError) as exc:
        print(f"error: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
