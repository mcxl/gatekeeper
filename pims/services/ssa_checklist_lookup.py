"""CCVS-keyed lookup over ``pims/audit_checklist.xlsx``.

Phase 1 helper for the SSA pipeline (workflow #1). Distinct from
``pims/services/checklist_matcher.py`` — that module operates on
Supabase ``checklist_items`` rows and returns severity states. This
module reads the static reviewer checklist xlsx and exposes a
``ChecklistMatch`` keyed by CCVS code, providing the verbatim values
the SSA pipeline writes into PIMS-Enriched / Staging xlsx fields.

The current ``audit_checklist.xlsx`` only carries
``Category, Criteria, Instruction`` columns. Any later columns
(``CCVS Code``, ``Legal Ref``, ``Action Description``,
``Recommendation``, ``Monitoring Note``) are picked up automatically
when the xlsx is extended; missing columns resolve to blank strings,
never None — caller writes them straight into a cell.

Lookup contract (per workflow plan §"Field Defaults"):

    ChecklistLookup.from_xlsx(path).match(ccvs_code) -> ChecklistMatch | None
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl

log = logging.getLogger(__name__)


@dataclass(frozen=True)
class ChecklistMatch:
    ccvs_code: str
    ccvs_category: str
    action_description: str
    recommendation: str
    legal_ref: str
    monitoring_note: str
    criteria: str = ""


# Header-name → field. Lowercased + whitespace-collapsed for matching.
# ``Instruction`` (the operational guidance column in the current xlsx)
# maps to ``action_description`` — that's what reviewers want in the
# action-register cell on the staging xlsx.
_HEADER_MAP = {
    "ccvs code": "ccvs_code",
    "ccvs_code": "ccvs_code",
    "category": "ccvs_category",
    "ccvs category": "ccvs_category",
    "ccvs_category": "ccvs_category",
    "action description": "action_description",
    "action_description": "action_description",
    "instruction": "action_description",
    "recommendation": "recommendation",
    "legal ref": "legal_ref",
    "legal_ref": "legal_ref",
    "monitoring note": "monitoring_note",
    "monitoring_note": "monitoring_note",
    "criteria": "criteria",
}

# Australian WHS legal-reference patterns embedded inside Criteria text,
# e.g. ``(WHS Reg cl.34-38)`` / ``(AS 1742.3)`` / ``(WHS Act s.19)``.
# Conservative — extracts only well-shaped citations, no fuzzy guesses.
_LEGAL_REF_RE = re.compile(
    r"\(("
    r"WHS\s+(?:Act|Reg)[^)]+|"
    r"AS\s*\d[\d./\s-]*|"
    r"AS/NZS\s*\d[\d./\s-]*"
    r")\)",
    re.IGNORECASE,
)


def _extract_legal_ref(criteria: str) -> str:
    if not criteria:
        return ""
    m = _LEGAL_REF_RE.search(criteria)
    return m.group(1).strip() if m else ""


# Token-set tools for the auto-matcher (match_observation). Stopword
# list is deliberately tight — drop only true noise words; keep WHS
# domain terms (eg. "first", "high", "edge") that carry signal.
_MATCHER_STOP = frozenset(
    "a an the is are was were be been being and or of to for in on at by "
    "with from as that this it not no any all each every some none has have "
    "had do did does we you they i me my our your their he she his her its "
    "but if then so than which who whom whose where when why how"
    .split()
)
_TOKEN_RE = re.compile(r"[a-z0-9]+")


def _tokens(text: str) -> set[str]:
    if not text:
        return set()
    return {
        w for w in _TOKEN_RE.findall(text.lower())
        if len(w) >= 3 and w not in _MATCHER_STOP
    }

# `01. Planning ...` / `01.02 ...` leading-number extraction. Used to
# synthesise a CCVS code when the xlsx doesn't carry one explicitly.
_LEADING_NUM = re.compile(r"^\s*(\d{1,3})(?:\.(\d{1,3}))?")


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _leading_num(value: object) -> tuple[str, str]:
    """Return (cat_num, item_num) extracted from a leading numeric prefix.

    ``"01. Planning and Risk Management"`` -> (``"01"``, ``""``).
    ``"02. Does the site sign include..."`` (criteria) -> (``"02"``, ``""``).
    Returns ``("", "")`` when nothing matches.
    """
    if value is None:
        return "", ""
    m = _LEADING_NUM.match(str(value))
    if not m:
        return "", ""
    return m.group(1), (m.group(2) or "")


def _synthesise_ccvs_code(cat_value: object, criteria_value: object) -> str:
    """Build ``"<cat>.<item>"`` from leading numbers on Category + Criteria.

    Returns ``""`` if either side has no numeric prefix — the caller then
    skips this row (no key, no match possible).
    """
    cat_num, _ = _leading_num(cat_value)
    item_num, _ = _leading_num(criteria_value)
    if not cat_num or not item_num:
        return ""
    return f"{int(cat_num):02d}.{int(item_num):02d}"


@dataclass
class ChecklistLookup:
    by_code: dict[str, ChecklistMatch]

    @classmethod
    def from_xlsx(
        cls,
        path: Path,
        sheet_names: Iterable[str] | None = None,
    ) -> "ChecklistLookup":
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = list(sheet_names) if sheet_names else list(wb.sheetnames)

        by_code: dict[str, ChecklistMatch] = {}
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                log.warning("checklist sheet %s missing from %s", sheet, path)
                continue
            ws = wb[sheet]
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                continue
            header_to_idx: dict[str, int] = {}
            for idx, cell in enumerate(header_row):
                key = _HEADER_MAP.get(_norm_header(cell))
                if key and key not in header_to_idx:
                    header_to_idx[key] = idx
            cat_idx = header_to_idx.get("ccvs_category")
            crit_idx = header_to_idx.get("criteria")
            # Heuristic fallback: first column = category, second = criteria
            # when the headers are literal "Category" / "Criteria".
            if cat_idx is None and len(header_row) > 0:
                cat_idx = 0
            if crit_idx is None and len(header_row) > 1:
                crit_idx = 1

            for row in rows:
                if not row or all((c is None or str(c).strip() == "") for c in row):
                    continue

                def cell(field: str) -> str:
                    i = header_to_idx.get(field)
                    if i is None or i >= len(row):
                        return ""
                    v = row[i]
                    return "" if v is None else str(v).strip()

                ccvs_code = cell("ccvs_code")
                if not ccvs_code:
                    cat_v = row[cat_idx] if cat_idx is not None and cat_idx < len(row) else ""
                    crit_v = row[crit_idx] if crit_idx is not None and crit_idx < len(row) else ""
                    ccvs_code = _synthesise_ccvs_code(cat_v, crit_v)
                if not ccvs_code:
                    continue

                cat_v = row[cat_idx] if cat_idx is not None and cat_idx < len(row) else ""
                crit_v = row[crit_idx] if crit_idx is not None and crit_idx < len(row) else ""
                criteria_text = cell("criteria") or (str(crit_v).strip() if crit_v else "")
                # Pull a legal_ref out of the criteria text when the
                # column itself is blank — current xlsx lacks the column
                # but criteria like ``(WHS Reg cl.34-38)`` carry it.
                legal_ref = cell("legal_ref") or _extract_legal_ref(criteria_text)
                match = ChecklistMatch(
                    ccvs_code=ccvs_code,
                    ccvs_category=cell("ccvs_category") or (str(cat_v).strip() if cat_v else ""),
                    action_description=cell("action_description"),
                    recommendation=cell("recommendation"),
                    legal_ref=legal_ref,
                    monitoring_note=cell("monitoring_note"),
                    criteria=criteria_text,
                )
                # First sheet wins on duplicate keys (the <$250K sheet ships
                # first; >$250K is the same coding).
                by_code.setdefault(ccvs_code.lower(), match)

        return cls(by_code=by_code)

    def match(self, ccvs_code: str) -> ChecklistMatch | None:
        if not ccvs_code:
            return None
        return self.by_code.get(ccvs_code.strip().lower())

    def match_observation(
        self,
        observation_text: str,
        min_overlap: int = 2,
        min_score: float = 0.40,
        min_margin: float = 0.10,
    ) -> ChecklistMatch | None:
        """Best-fit checklist row for a free-form observation.

        Scoring is token-recall against the observation:
            score = |obs_tokens ∩ candidate_tokens| / |obs_tokens|

        Conservative gate — a candidate wins iff ALL hold:
          - ``overlap`` (the size of the token intersection) ≥ ``min_overlap``
          - ``score`` ≥ ``min_score``
          - ``score`` exceeds the second-best candidate's score by at
            least ``min_margin`` (an unambiguous winner)

        Otherwise returns ``None`` — the caller should leave the row at
        ``status="Unmatched"``. Defaults were probed against the v1
        ``audit_checklist.xlsx`` and ten representative audit
        observations: they accept the obvious wins and reject the ties.
        """
        obs_tokens = _tokens(observation_text)
        if len(obs_tokens) < min_overlap:
            return None

        scored: list[tuple[float, int, ChecklistMatch]] = []
        for m in self.by_code.values():
            cand_tokens = _tokens(f"{m.criteria} {m.ccvs_category}")
            if not cand_tokens:
                continue
            overlap = len(obs_tokens & cand_tokens)
            if overlap < min_overlap:
                continue
            score = overlap / len(obs_tokens)
            if score < min_score:
                continue
            scored.append((score, overlap, m))

        if not scored:
            return None
        # Highest score wins; ties broken by raw overlap count, then by
        # CCVS code (stable / deterministic).
        scored.sort(key=lambda t: (-t[0], -t[1], t[2].ccvs_code))
        top = scored[0]
        if len(scored) >= 2 and (top[0] - scored[1][0]) < min_margin:
            return None
        return top[2]
