"""Controlled photo backfill for the requested July 2026 audits.

Only rows for the requested site/date pairs are considered. A photo is linked
only when the source audit register identifies the same finding and photo
reference. Existing photo links are never overwritten.
"""
from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\AlanRichardson\OneDrive - AuditCo\RPD\SSA")
CONTROLLED = Path(
    r"C:\Users\AlanRichardson\Documents\agentic-os-workspace\audit-concierge\controlled\pims-photo-backfill-2026-08-31"
)
BUCKET = "pims-photos"


AUDITS = {
    "chelmsford-botany-260721": {
        "site": "25 Chelmsford Ave Botany",
        "date": "2026-07-21",
        "folder": "SSA - 25 Chelmsford Ave Botany 260721",
        "version": "SSA - 260721 -",
        "kind": "reviewed-json",
    },
    "shackle-brookvale-260715": {
        "site": "21 Shackle Avenue Brookvale",
        "date": "2026-07-15",
        "folder": "SSA - 21 Shackle Avenue Brookvale 260715",
        "version": "SSA - 260715 -",
        "kind": "workbook",
    },
    "wentworth-liberty-grove-260730": {
        "site": "8 Wentworth Dr Liberty Grove",
        "date": "2026-07-30",
        "folder": "SSA - 8 Wentworth Dr Liberty Grove 260730",
        "version": "SSA - 260730-V2",
        "kind": "explicit-report-map",
    },
}


RULES = [
    # The source workbook has a photo reference for each of these findings.
    {"audit": "shackle-brookvale-260715", "keywords": ["project site sign", "not displayed"], "source_id": "RPD-SSA-260715-02-0011"},
    {"audit": "shackle-brookvale-260715", "keywords": ["induction qr code", "not displayed"], "source_id": "RPD-SSA-260715-02-0017"},
    {"audit": "shackle-brookvale-260715", "keywords": ["emergency contact details", "nearest medical facility"], "source_id": "RPD-SSA-260715-02-0012"},
    {"audit": "shackle-brookvale-260715", "keywords": ["non-standard ladder-beam scaffold", "engineering design"], "source_id": "RPD-SSA-260715-02-0013"},
    {"audit": "shackle-brookvale-260715", "keywords": ["extension leads", "scaffold deck", "test tags"], "source_id": "RPD-SSA-260715-02-0014"},
    # The report appendix identifies these photo numbers. The files are the
    # matching source images in the same audit folder.
    {"audit": "wentworth-liberty-grove-260730", "section": "1.14", "keywords": ["no task-specific swms", "fall risk exceeding 2 m"], "photo_number": 20, "filename": "IMG_8257.jpg"},
    {"audit": "wentworth-liberty-grove-260730", "section": "6.1", "keywords": ["no task-specific swms was available"], "photo_number": 12, "filename": "IMG_8249.PNG"},
    {"audit": "wentworth-liberty-grove-260730", "section": "6.8", "keywords": ["plant and equipment register", "not available"], "photo_number": 19, "filename": "IMG_8256.jpg"},
    {"audit": "wentworth-liberty-grove-260730", "section": "6.13", "keywords": ["not readily accessible as auditable records"], "photo_number": 14, "filename": "IMG_8251.PNG"},
    {"audit": "wentworth-liberty-grove-260730", "section": "7.1", "keywords": ["charging equipment", "inspection and test tag"], "photo_number": 18, "filename": "IMG_8255.jpg"},
]


def norm(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def load_env() -> tuple[str, str]:
    env_file = Path(r"C:\Users\AlanRichardson\Documents\agentic-os-workspace\audit-concierge\.env")
    if env_file.exists():
        for raw in env_file.read_text(encoding="utf-8").splitlines():
            if not raw.strip() or raw.lstrip().startswith("#") or "=" not in raw:
                continue
            name, value = raw.split("=", 1)
            os.environ.setdefault(name.strip(), value.strip().strip('"').strip("'"))
    url = (os.environ.get("RPD_SUPABASE_URL") or "").rstrip("/")
    key = os.environ.get("RPD_SUPABASE_SERVICE_KEY") or ""
    if not url or not key:
        raise RuntimeError("RPD Supabase URL/service key is not available")
    return url, key


def rest_headers(key: str) -> dict[str, str]:
    return {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json", "Prefer": "return=minimal"}


def source_root(audit: dict[str, Any]) -> Path:
    return ROOT / audit["folder"]


def load_workbook_records(audit_key: str) -> dict[str, dict[str, Any]]:
    audit = AUDITS[audit_key]
    root = source_root(audit)
    files = root / f'{audit["version"]} files'
    # Chelmsford uses a report-named REVIEW workbook in this folder.
    if audit_key == "chelmsford-botany-260721":
        books = list(files.glob("*REVIEW.xlsx"))
    else:
        books = [files / "REVIEW.xlsx"]
    if not books or not books[0].exists():
        return {}
    ws = load_workbook(books[0], read_only=True, data_only=True)["Observation"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "") for h in rows[0]]
    output = {}
    for values in rows[1:]:
        record = dict(zip(headers, values))
        if record.get("id"):
            output[str(record["id"])] = record
    return output


def load_json_records(audit_key: str) -> list[dict[str, Any]]:
    audit = AUDITS[audit_key]
    root = source_root(audit)
    files = root / f'{audit["version"]} files'
    json_file = files / "reviewed-observations.json"
    return json.loads(json_file.read_text(encoding="utf-8")) if json_file.exists() else []


def source_for_rule(rule: dict[str, Any]) -> dict[str, Any] | None:
    audit = AUDITS[rule["audit"]]
    root = source_root(audit)
    if rule["audit"] == "shackle-brookvale-260715":
        records = load_workbook_records(rule["audit"])
        record = records.get(rule["source_id"])
        if not record or not record.get("photo refs") or not record.get("storage path"):
            return None
        photo_ref = str(record["photo refs"]).split(",", 1)[0].strip()
        local = (root / f'{audit["version"]} photos' / str(record["storage path"])).resolve()
        return {"photo_ref": photo_ref, "source_id": rule["source_id"], "local": local}
    if rule["audit"] == "wentworth-liberty-grove-260730":
        local = (root / f'{audit["version"]} - photos' / "image" / rule["filename"]).resolve()
        return {"photo_ref": f"Photo {rule['photo_number']}", "source_id": f"report-photo-{rule['photo_number']}", "local": local}
    return None


def get_rows(session: requests.Session, url: str, key: str) -> list[dict[str, Any]]:
    sites = ",".join(a["site"] for a in AUDITS.values())
    # Fetch all observations for the three sites, then apply normalized matching
    # locally so punctuation differences do not affect the scope.
    response = session.get(
        f"{url}/rest/v1/pims_observations",
        headers=rest_headers(key),
        params={"select": "id,site_address,observation_date,observation_text,photo_url,photo_refs,filename,source_pdf", "or": "(" + ",".join(f"site_address.ilike.*{a['site'].split()[0]}*" for a in AUDITS.values()) + ")"},
        timeout=60,
    )
    response.raise_for_status()
    return [r for r in response.json() if not r.get("photo_url") and any(norm(r.get("site_address")) == norm(a["site"]) and r.get("observation_date") == a["date"] for a in AUDITS.values())]


def count_rows(session: requests.Session, url: str, key: str, table: str) -> int:
    response = session.get(f"{url}/rest/v1/{table}", headers={**rest_headers(key), "Prefer": "count=exact"}, params={"select": "id", "limit": "1"}, timeout=60)
    response.raise_for_status()
    return int(response.headers.get("content-range", "0/0").split("/")[-1])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    url, key = load_env()
    session = requests.Session()
    before = {"pims_observations": count_rows(session, url, key, "pims_observations"), "pims_staging": count_rows(session, url, key, "pims_staging")}
    rows = get_rows(session, url, key)
    manifest: dict[str, Any] = {"generated_at": datetime.now(timezone.utc).isoformat(), "mode": "apply" if args.apply else "dry-run", "requested_audits": AUDITS, "before_counts": before, "matched": [], "unmatched": [], "applied": []}
    for row in rows:
        text = (row.get("observation_text") or "").lower()
        matching = []
        for rule in RULES:
            audit = AUDITS[rule["audit"]]
            if norm(row.get("site_address")) != norm(audit["site"]) or row.get("observation_date") != audit["date"]:
                continue
            section = (row.get("observation_text") or "").split(" -", 1)[0].strip()
            if rule.get("section") and section != rule["section"]:
                continue
            if all(k in text for k in rule["keywords"]):
                matching.append(rule)
        if len(matching) != 1:
            manifest["unmatched"].append({"row_id": row["id"], "site_address": row.get("site_address"), "observation_text": row.get("observation_text"), "reason": f"rule count {len(matching)}"})
            continue
        source = source_for_rule(matching[0])
        if not source or not source["local"].exists():
            manifest["unmatched"].append({"row_id": row["id"], "reason": "source photo is not available"})
            continue
        item = {"row_id": row["id"], "site_address": row.get("site_address"), "observation_date": row.get("observation_date"), "photo_ref": source["photo_ref"], "source_id": source["source_id"], "local_path": str(source["local"]), "storage_path": f"RPD-SSA/historical/{matching[0]['audit']}/{source['local'].name}"}
        manifest["matched"].append(item)
    if args.apply:
        for item in manifest["matched"]:
            local = Path(item["local_path"])
            storage = item["storage_path"]
            content_type = mimetypes.guess_type(local.name)[0] or "application/octet-stream"
            upload = session.put(f"{url}/storage/v1/object/{BUCKET}/{storage}", headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": content_type, "x-upsert": "true"}, data=local.read_bytes(), timeout=120)
            upload.raise_for_status()
            public_url = f"{url}/storage/v1/object/public/{BUCKET}/{storage}"
            patch = session.patch(f"{url}/rest/v1/pims_observations", headers=rest_headers(key), params={"id": f"eq.{item['row_id']}"}, json={"photo_url": public_url, "photo_refs": item["photo_ref"], "filename": local.name}, timeout=60)
            patch.raise_for_status()
            manifest["applied"].append({**item, "public_url": public_url})
    after = {"pims_observations": count_rows(session, url, key, "pims_observations"), "pims_staging": count_rows(session, url, key, "pims_staging")}
    manifest["after_counts"] = after
    manifest["counts_unchanged"] = before == after
    CONTROLLED.mkdir(parents=True, exist_ok=True)
    output = CONTROLLED / ("manifest-requested-apply.json" if args.apply else "manifest-requested-dry-run.json")
    output.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"mode": manifest["mode"], "candidate_rows": len(rows), "matched": len(manifest["matched"]), "unmatched": len(manifest["unmatched"]), "applied": len(manifest["applied"]), "counts_unchanged": manifest["counts_unchanged"], "manifest": str(output)}))
    for item in manifest["unmatched"]:
        print(f"UNMATCHED {item['row_id']}: {item['reason']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
