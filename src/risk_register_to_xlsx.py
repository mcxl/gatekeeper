#!/usr/bin/env python3
"""Convert the risk register data to a formatted Excel workbook.

Produces a two-sheet workbook:
  Sheet 1 — "Risk Register" (main table, summaries, hold points)
  Sheet 2 — "Matrix & Lists" (risk matrix, dropdown lists, definitions)

Styling uses the same colour scheme as the docx generator:
  Critical/High = FF0000, Medium = FFFF00, Low = 00FF00
  Header row = DBE5F1, font = Arial throughout.
"""

import sys

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# Import the default config from the docx module (used when run standalone)
from src.risk_register_to_docx import DEFAULT_CONFIG as _WATERLOO_CONFIG

# ── Constants (matching docx_style_standard) ──────────────────────
FONT_NAME = "Arial"
HEADER_BG = "DBE5F1"
RISK_COLOURS = {
    "Critical": "FF0000",
    "High":     "FF0000",
    "Medium":   "FFFF00",
    "Low":      "00FF00",
}
RISK_FONT_COLOURS = {
    "Critical": "FFFFFF",
    "High":     "FFFFFF",
    "Medium":   "000000",
    "Low":      "000000",
}
ALT_ROW_BG = "F2F2F2"
BLACK = "000000"
WHITE = "FFFFFF"

# ── Reusable styles ──────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FONT = Font(name=FONT_NAME, size=8, bold=True, color=BLACK)
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name=FONT_NAME, size=8, color=BLACK)
BODY_FONT_BOLD = Font(name=FONT_NAME, size=8, bold=True, color=BLACK)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

ALT_FILL = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")

# ── Risk matrix data ─────────────────────────────────────────────
LIKELIHOOD_CODES = ["A", "B", "C", "D", "E"]
LIKELIHOOD_LABELS = [
    "A — Almost Certain",
    "B — Likely",
    "C — Possible",
    "D — Unlikely",
    "E — Rare",
]
CONSEQUENCE_CODES = ["1", "2", "3"]
CONSEQUENCE_LABELS = ["1 — Minor", "2 — Moderate", "3 — Major"]

# Matrix[likelihood_index][consequence_index] = (level, score)
RISK_MATRIX = [
    [("High", 3),     ("Critical", 5), ("Critical", 6)],   # A
    [("Medium", 2),   ("High", 4),     ("Critical", 5)],   # B
    [("Low", 1),      ("Medium", 3),   ("High", 4)],       # C
    [("Low", 1),      ("Low", 2),      ("Medium", 3)],     # D
    [("Low", 1),      ("Low", 1),      ("Low", 2)],        # E
]

GATEKEEPER_CODES = ["WAH", "SIL", "ENV", "STR", "ASB", "LED", "TRF", "CHM", "WAT", "EMR"]
RISK_LEVELS = ["Critical (6)", "Critical (5)", "High (4)", "High (3)", "Medium (3)", "Medium (2)", "Low (2)", "Low (1)"]



def _risk_fill(level: str) -> PatternFill:
    """Return a PatternFill for a risk level."""
    bg = RISK_COLOURS.get(level, RISK_COLOURS["Low"])
    return PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def _risk_font(level: str) -> Font:
    """Return a Font with contrast colour for a risk level."""
    fc = RISK_FONT_COLOURS.get(level, BLACK)
    return Font(name=FONT_NAME, size=8, bold=True, color=fc)


def _extract_level(rating_str: str) -> str:
    """Extract the risk level word from a rating string like 'Critical (5)'."""
    for level in ("Critical", "High", "Medium", "Low"):
        if level in rating_str:
            return level
    return "Low"


def _apply_header(ws, row: int, col: int, text: str) -> None:
    """Apply header styling to a cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER


def _apply_risk_cell(ws, row: int, col: int, text: str) -> None:
    """Apply risk-colour styling to a cell."""
    cell = ws.cell(row=row, column=col, value=text)
    level = _extract_level(text)
    cell.font = _risk_font(level)
    cell.fill = _risk_fill(level)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _apply_body_cell(ws, row: int, col: int, text: str,
                     bold: bool = False, center: bool = False) -> None:
    """Apply standard body styling to a cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BODY_FONT_BOLD if bold else BODY_FONT
    cell.alignment = CENTER_ALIGN if center else BODY_ALIGN
    cell.border = THIN_BORDER


def build_sheet1_risk_register(ws, config: dict) -> None:
    """Build the Risk Register sheet with main table, summaries, and hold points."""
    ws.title = "Risk Register"
    risks = config["risks"]

    # ── Project header ────────────────────────────────────────────
    project_details = [
        ("Project:", config["project_name"]),
        ("PCBU / Principal Contractor:", config["pcbu"]),
        ("Jurisdiction:", config["jurisdiction"]),
        ("Date Prepared:", config["date"]),
        ("Prepared by:", config["prepared_by"]),
    ]
    for i, (label, value) in enumerate(project_details):
        label_cell = ws.cell(row=i + 1, column=1, value=label)
        label_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=BLACK)
        label_cell.alignment = Alignment(vertical="top")
        value_cell = ws.cell(row=i + 1, column=2, value=value)
        value_cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
        value_cell.alignment = Alignment(vertical="top")

    # ── Column headers (row 7) ────────────────────────────────────
    header_row = 7
    headers = [
        "#", "Task", "Code", "Hazard", "Likelihood\n(Pre)",
        "Consequence\n(Pre)", "Risk Rating\n(Pre-Controls)",
        "Controls", "Residual\nRisk", "Responsible\nPerson",
    ]
    for j, h in enumerate(headers):
        _apply_header(ws, header_row, j + 1, h)

    # ── Column widths (approximate cm → Excel character widths) ───
    col_widths = [4, 40, 6, 40, 12, 12, 14, 80, 12, 22]
    for j, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w

    # ── Data rows ─────────────────────────────────────────────────
    data_start = header_row + 1
    for i, risk in enumerate(risks):
        r = data_start + i
        _apply_body_cell(ws, r, 1, risk["no"], center=True)
        _apply_body_cell(ws, r, 2, risk["task"])
        _apply_body_cell(ws, r, 3, risk["code"], bold=True, center=True)
        _apply_body_cell(ws, r, 4, risk["hazard"])
        _apply_body_cell(ws, r, 5, risk["likelihood_pre"], center=True)
        _apply_body_cell(ws, r, 6, risk["consequence_pre"], center=True)
        _apply_risk_cell(ws, r, 7, risk["risk_pre"])
        _apply_body_cell(ws, r, 8, risk["controls"])
        _apply_risk_cell(ws, r, 9, risk["residual_risk"])
        _apply_body_cell(ws, r, 10, risk["responsible"])

        # Alternate row shading (skip risk-coloured cells 7 and 9)
        if i % 2 == 1:
            alt = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
            for col in [1, 2, 3, 4, 5, 6, 8, 10]:
                ws.cell(row=r, column=col).fill = alt

    data_end = data_start + len(risks) - 1

    # ── Data validation — Likelihood dropdown (column 5) ──────────
    likelihood_formula = '"A — Almost Certain,B — Likely,C — Possible,D — Unlikely,E — Rare"'
    dv_likelihood = DataValidation(
        type="list", formula1=likelihood_formula, allow_blank=True,
    )
    dv_likelihood.error = "Select a valid likelihood (A–E)"
    dv_likelihood.errorTitle = "Invalid Likelihood"
    dv_likelihood.prompt = "Select likelihood A–E"
    dv_likelihood.promptTitle = "Likelihood"
    ws.add_data_validation(dv_likelihood)
    dv_likelihood.add(f"E{data_start}:E{data_end}")

    # ── Data validation — Consequence dropdown (column 6) ─────────
    consequence_formula = '"1 — Minor,2 — Moderate,3 — Major"'
    dv_consequence = DataValidation(
        type="list", formula1=consequence_formula, allow_blank=True,
    )
    dv_consequence.error = "Select a valid consequence (1–3)"
    dv_consequence.errorTitle = "Invalid Consequence"
    dv_consequence.prompt = "Select consequence 1–3"
    dv_consequence.promptTitle = "Consequence"
    ws.add_data_validation(dv_consequence)
    dv_consequence.add(f"F{data_start}:F{data_end}")

    # ── Data validation — Code dropdown (column 3) ────────────────
    code_formula = '"' + ",".join(GATEKEEPER_CODES) + '"'
    dv_code = DataValidation(
        type="list", formula1=code_formula, allow_blank=True,
    )
    dv_code.error = "Select a valid Gatekeeper code"
    dv_code.errorTitle = "Invalid Code"
    ws.add_data_validation(dv_code)
    dv_code.add(f"C{data_start}:C{data_end}")

    # ── Data validation — Residual Risk dropdown (column 9) ───────
    residual_formula = '"' + ",".join(RISK_LEVELS) + '"'
    dv_residual = DataValidation(
        type="list", formula1=residual_formula, allow_blank=True,
    )
    dv_residual.error = "Select a valid residual risk rating"
    dv_residual.errorTitle = "Invalid Residual Risk"
    ws.add_data_validation(dv_residual)
    dv_residual.add(f"I{data_start}:I{data_end}")

    # ── Risk Rating (column 7) — write the pre-calculated value ─
    # Static values from the risks data ensure text is always visible.
    # Conditional formatting (below) applies the colour coding.
    # For any new rows added by the user, an XLOOKUP formula template
    # is provided on the 'Matrix & Lists' sheet.
    for i, risk in enumerate(risks):
        r = data_start + i
        _apply_risk_cell(ws, r, 7, risk["risk_pre"])

    # ── Conditional formatting for Risk Rating (column 7) ─────────
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    critical_font = Font(name=FONT_NAME, size=8, bold=True, color=WHITE)
    high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    high_font = Font(name=FONT_NAME, size=8, bold=True, color=WHITE)
    medium_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    medium_font = Font(name=FONT_NAME, size=8, bold=True, color=BLACK)
    low_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    low_font = Font(name=FONT_NAME, size=8, bold=True, color=BLACK)

    # Apply conditional formatting to both Risk Rating (G) and Residual Risk (I)
    for col_letter in ["G", "I"]:
        cell_range = f"{col_letter}{data_start}:{col_letter}{data_end}"
        ref = f"{col_letter}{data_start}"
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("Critical",{ref})))'],
                fill=critical_fill, font=critical_font,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("High",{ref})))'],
                fill=high_fill, font=high_font,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("Medium",{ref})))'],
                fill=medium_fill, font=medium_font,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("Low",{ref})))'],
                fill=low_fill, font=low_font,
            ),
        )

    # ── Summary section ───────────────────────────────────────────
    summary_start = data_end + 3
    ws.cell(row=summary_start, column=1, value="Risk Profile Summary").font = Font(
        name=FONT_NAME, size=12, bold=True, color=BLACK,
    )

    # Pre-controls summary
    pre_row = summary_start + 2
    ws.cell(row=pre_row, column=1, value="Pre-Controls").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    pre_row += 1
    _apply_header(ws, pre_row, 1, "Risk Rating")
    _apply_header(ws, pre_row, 2, "Count")

    pre_data = config["pre_summary"]
    for i, (rating, count) in enumerate(pre_data):
        r = pre_row + 1 + i
        level = _extract_level(rating)
        cell_rating = ws.cell(row=r, column=1, value=rating)
        cell_rating.font = _risk_font(level)
        cell_rating.fill = _risk_fill(level)
        cell_rating.border = THIN_BORDER
        cell_count = ws.cell(row=r, column=2, value=int(count))
        cell_count.font = BODY_FONT
        cell_count.alignment = CENTER_ALIGN
        cell_count.border = THIN_BORDER

    # Post-controls summary
    post_row = pre_row + len(pre_data) + 2
    ws.cell(row=post_row, column=1, value="Post-Controls (Residual)").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    post_row += 1
    _apply_header(ws, post_row, 1, "Risk Rating")
    _apply_header(ws, post_row, 2, "Count")

    post_data = config["post_summary"]
    for i, (rating, count) in enumerate(post_data):
        r = post_row + 1 + i
        level = _extract_level(rating)
        cell_rating = ws.cell(row=r, column=1, value=rating)
        cell_rating.font = _risk_font(level)
        cell_rating.fill = _risk_fill(level)
        cell_rating.border = THIN_BORDER
        cell_count = ws.cell(row=r, column=2, value=int(count))
        cell_count.font = BODY_FONT
        cell_count.alignment = CENTER_ALIGN
        cell_count.border = THIN_BORDER

    # ── Critical Hold Points ──────────────────────────────────────
    hp_row = post_row + len(config["post_summary"]) + 3
    ws.cell(row=hp_row, column=1, value="Critical Hold Points").font = Font(
        name=FONT_NAME, size=12, bold=True, color=BLACK,
    )
    hp_row += 1
    for i, hp in enumerate(config["hold_points"]):
        cell = ws.cell(row=hp_row + i, column=1, value=f"• {hp}")
        cell.font = Font(name=FONT_NAME, size=9, color=BLACK)
        cell.alignment = Alignment(wrap_text=True)
        # Merge across columns for readability
        ws.merge_cells(
            start_row=hp_row + i, start_column=1,
            end_row=hp_row + i, end_column=6,
        )

    # ── References ────────────────────────────────────────────────
    ref_row = hp_row + len(config["hold_points"]) + 2
    ws.cell(row=ref_row, column=1, value="References").font = Font(
        name=FONT_NAME, size=12, bold=True, color=BLACK,
    )
    ref_row += 1
    for i, ref in enumerate(config["references"]):
        cell = ws.cell(row=ref_row + i, column=1, value=f"• {ref}")
        cell.font = Font(name=FONT_NAME, size=8, color=BLACK)
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(
            start_row=ref_row + i, start_column=1,
            end_row=ref_row + i, end_column=6,
        )

    # ── Freeze panes and auto-filter ──────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:J{data_end}"

    # ── Print setup: landscape, fit to 1 page wide, repeat header ─
    ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # as many pages tall as needed
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Set row heights for data rows
    for i in range(len(risks)):
        ws.row_dimensions[data_start + i].height = 80


def build_sheet2_matrix(ws) -> None:
    """Build the Matrix & Lists sheet with risk matrix, dropdown lists, and definitions."""
    ws.title = "Matrix & Lists"

    # ── Risk Matrix ───────────────────────────────────────────────
    ws.cell(row=1, column=1, value="").font = Font(name=FONT_NAME, size=8, color=BLACK)

    # Consequence headers in B1:D1 (single character for XLOOKUP key)
    for j, code in enumerate(CONSEQUENCE_CODES):
        cell = ws.cell(row=1, column=j + 2, value=code)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Full consequence labels in row 2 for display
    _apply_header(ws, 2, 1, "Likelihood \\ Consequence")
    for j, label in enumerate(CONSEQUENCE_LABELS):
        _apply_header(ws, 2, j + 2, label)

    # Likelihood rows with matrix values
    for i, (code, label) in enumerate(zip(LIKELIHOOD_CODES, LIKELIHOOD_LABELS)):
        r = 3 + i
        # Column A: likelihood code (single letter for XLOOKUP key)
        key_cell = ws.cell(row=r, column=1, value=code)
        key_cell.font = BODY_FONT_BOLD
        key_cell.alignment = CENTER_ALIGN
        key_cell.border = THIN_BORDER

        for j in range(3):
            level, score = RISK_MATRIX[i][j]
            text = f"{level} ({score})"
            _apply_risk_cell(ws, r, j + 2, text)

    # Full likelihood labels in column E for reference
    _apply_header(ws, 2, 5, "Likelihood Label")
    for i, label in enumerate(LIKELIHOOD_LABELS):
        cell = ws.cell(row=3 + i, column=5, value=label)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 50

    # ── Definitions section ───────────────────────────────────────
    def_start = 10
    ws.cell(row=def_start, column=1, value="Likelihood Definitions").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    likelihood_defs = [
        ("A — Almost Certain", "Expected to occur in most circumstances"),
        ("B — Likely", "Will probably occur in most circumstances"),
        ("C — Possible", "Might occur at some time"),
        ("D — Unlikely", "Could occur but not expected"),
        ("E — Rare", "May occur only in exceptional circumstances"),
    ]
    _apply_header(ws, def_start + 1, 1, "Level")
    _apply_header(ws, def_start + 1, 2, "Description")
    for i, (level, desc) in enumerate(likelihood_defs):
        r = def_start + 2 + i
        _apply_body_cell(ws, r, 1, level, bold=True)
        _apply_body_cell(ws, r, 2, desc)

    cons_start = def_start + len(likelihood_defs) + 4
    ws.cell(row=cons_start, column=1, value="Consequence Definitions").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    consequence_defs = [
        ("1 — Minor", "First aid treatment; minor property damage"),
        ("2 — Moderate", "Medical treatment; significant property damage"),
        ("3 — Major", "Fatality, permanent disability, or major structural failure"),
    ]
    _apply_header(ws, cons_start + 1, 1, "Level")
    _apply_header(ws, cons_start + 1, 2, "Description")
    for i, (level, desc) in enumerate(consequence_defs):
        r = cons_start + 2 + i
        _apply_body_cell(ws, r, 1, level, bold=True)
        _apply_body_cell(ws, r, 2, desc)

    # ── Gatekeeper Code List ──────────────────────────────────────
    code_start = cons_start + len(consequence_defs) + 4
    ws.cell(row=code_start, column=1, value="Gatekeeper Hazard Codes").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    code_defs = [
        ("WAH", "Work at height — collective height access (EWP, scaffold, ladder)"),
        ("SIL", "Silica and dust"),
        ("ENV", "Environmental and chemical"),
        ("STR", "Structural"),
        ("ASB", "Asbestos"),
        ("LED", "Lead"),
        ("TRF", "Traffic and public interface"),
        ("CHM", "Chemical hazards"),
        ("WAT", "Water / waterproofing hazards"),
        ("EMR", "Emergency response"),
    ]
    _apply_header(ws, code_start + 1, 1, "Code")
    _apply_header(ws, code_start + 1, 2, "Category")
    for i, (code, desc) in enumerate(code_defs):
        r = code_start + 2 + i
        _apply_body_cell(ws, r, 1, code, bold=True, center=True)
        _apply_body_cell(ws, r, 2, desc)

    # ── Risk Level List (for dropdown reference) ──────────────────
    level_start = code_start + len(code_defs) + 4
    ws.cell(row=level_start, column=1, value="Risk Levels").font = Font(
        name=FONT_NAME, size=10, bold=True, color=BLACK,
    )
    _apply_header(ws, level_start + 1, 1, "Rating")
    _apply_header(ws, level_start + 1, 2, "Action Required")
    level_actions = [
        ("Critical (6)", "Immediate stop work — controls must be verified and approved before proceeding"),
        ("Critical (5)", "Immediate stop work — controls must be verified and approved before proceeding"),
        ("High (4)", "Senior management attention required — additional controls needed"),
        ("High (3)", "Senior management attention required — additional controls needed"),
        ("Medium (3)", "Manage with specific procedures and monitoring"),
        ("Medium (2)", "Manage with specific procedures and monitoring"),
        ("Low (2)", "Manage by routine procedures"),
        ("Low (1)", "Manage by routine procedures"),
    ]
    for i, (rating, action) in enumerate(level_actions):
        r = level_start + 2 + i
        level = _extract_level(rating)
        cell_rating = ws.cell(row=r, column=1, value=rating)
        cell_rating.font = _risk_font(level)
        cell_rating.fill = _risk_fill(level)
        cell_rating.border = THIN_BORDER
        cell_rating.alignment = CENTER_ALIGN
        _apply_body_cell(ws, r, 2, action)

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def build_workbook(config: dict | None = None) -> Workbook:
    """Build the complete two-sheet workbook."""
    if config is None:
        config = _WATERLOO_CONFIG
    wb = Workbook()
    ws1 = wb.active
    build_sheet1_risk_register(ws1, config)

    ws2 = wb.create_sheet()
    build_sheet2_matrix(ws2)

    return wb


if __name__ == "__main__":
    output_path = "output/Risk_Register_18_Danks_St_Waterloo.xlsx"
    if len(sys.argv) > 1:
        output_path = sys.argv[1]

    wb = build_workbook()
    wb.save(output_path)
    print(f"Risk register (Excel) saved to {output_path}")
