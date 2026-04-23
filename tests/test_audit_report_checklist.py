"""Tests for checklist loading (sheet selection by project_value) and matching."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from pims import audit_report_docx as arpt


def _make_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    # Default sheet removed
    wb.remove(wb.active)
    high = wb.create_sheet(arpt.SHEET_HIGH)
    high.append(["Category", "Criteria", "Instruction", "ccvs_code", "ccvs_category",
                 "observation_text_enriched"])
    high.append(["HIGH", "Do the thing (>=250K)", "Check the thing", "WAH-H6",
                 "Working at Height", "Worker at height"])
    low = wb.create_sheet(arpt.SHEET_LOW)
    low.append(["Category", "Criteria", "Instruction", "ccvs_code", "ccvs_category",
                "observation_text_enriched"])
    low.append(["LOW", "Do the small thing (<250K)", "Check the small thing", "SYS-L1",
                "Systems", "Small site check"])
    wb.save(path)


@pytest.fixture
def xlsx(tmp_path):
    p = tmp_path / "checklist.xlsx"
    _make_xlsx(p)
    return p


def test_sheet_selection_boundary_at_250k(xlsx):
    # Exactly 250000 -> high sheet
    rows = arpt.load_checklist(250000, xlsx)
    assert len(rows) == 1
    assert rows[0].category == "HIGH"

    # 249999.99 -> low sheet
    rows = arpt.load_checklist(249999.99, xlsx)
    assert rows[0].category == "LOW"

    # 250001 -> high
    rows = arpt.load_checklist(250001, xlsx)
    assert rows[0].category == "HIGH"


def test_load_missing_xlsx_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        arpt.load_checklist(100000, tmp_path / "missing.xlsx")


def test_load_null_project_value_raises(xlsx):
    with pytest.raises(ValueError):
        arpt.load_checklist(None, xlsx)


def test_match_by_ccvs_code(xlsx):
    rows = arpt.load_checklist(300000, xlsx)
    obs = {"ccvs_code": "WAH-H6", "observation_text": "worker without harness"}
    matched, ratio = arpt.match_observation(obs, rows)
    assert matched is rows[0]
    assert ratio == 1.0


def test_match_fallback_ratio(xlsx):
    rows = arpt.load_checklist(300000, xlsx)
    # no ccvs_code, but text strongly overlaps the row blob
    obs = {
        "ccvs_category": "HIGH",
        "observation_text_enriched": "Do the thing (>=250K)",
    }
    matched, ratio = arpt.match_observation(obs, rows)
    assert matched is rows[0]
    assert ratio >= 0.75


def test_no_match_when_unrelated(xlsx):
    rows = arpt.load_checklist(300000, xlsx)
    obs = {"observation_text_enriched": "completely unrelated nonsense xyzzy"}
    matched, _ = arpt.match_observation(obs, rows)
    assert matched is None
